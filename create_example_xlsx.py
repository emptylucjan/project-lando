"""Create an example produkty.xlsx file with test product URLs."""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Produkty"

# Header
ws["A1"] = "URL"
ws["B1"] = "Max Cena"
ws["A1"].font = openpyxl.styles.Font(bold=True)
ws["B1"].font = openpyxl.styles.Font(bold=True)
ws.column_dimensions["A"].width = 100
ws.column_dimensions["B"].width = 15

# Example products
products = [
    ("https://www.zalando.pl/nike-sportswear-air-force-1-07-sneakersy-niskie-white-ni112n022-a11.html", 500),
    ("https://www.zalando.pl/nike-sportswear-air-force-1-07-sneakersy-niskie-black-ni112n022-q11.html", 450),
]

for url, max_price in products:
    ws.append([url, max_price])

filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "produkty.xlsx")
wb.save(filepath)
print(f"Utworzono: {filepath}")
print(f"Produktów: {len(products)}")
