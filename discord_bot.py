"""
Discord Bot — Zalando Monitor Controller

Uruchomienie: python discord_bot.py

Komendy:
  !scan [keyword]         — skanuj konkretny keyword (lub pierwszy z listy)
  !scanall                — skanuje wszystkie keywordy po kolei
  !keywords               — pokaż listę keywordów z filtrami
  !add <kw> [eur:X] [disc:X] — dodaj keyword z opcjonalnym filtrem EUR i/lub % rabatu
  !remove <keyword>       — usuń keyword z listy
  !prog <kwota>           — zmień domyślny próg buy price w EUR
  !status                 — aktualny config
  !help                   — lista komend

Reakcje na kartach produktów:
  ✅  — pobierz rozmiary i dodaj do xlsx
  ❌  — odrzuć i usuń kartę

Składnia !add:
  !add shox               → domyślny próg EUR (z !prog)
  !add shox eur:80        → alert gdy buy price ≤ 80 EUR
  !add shox disc:40       → alert gdy rabat ≥ 40%
  !add jordan eur:90 disc:35  → alert gdy jedno z dwóch spełnione
"""

import asyncio
import datetime
import io
import json
import logging
import os
import re
import sys
import time

import discord
from discord.ext import commands, tasks

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_ROOT)

# ─────────────────────────────────────────────
_cfg_path = os.path.join(PROJECT_ROOT, "bot_config.json")
with open(_cfg_path, "r", encoding="utf-8") as _f:
    _bcfg = json.load(_f)
BOT_TOKEN   = _bcfg["bot_token"]
CHANNEL_ID  = _bcfg.get("channel_id", 1115326621562450043)
XLSX_FILE   = os.path.join(PROJECT_ROOT, "produkty.xlsx")
CONFIG_FILE = os.path.join(PROJECT_ROOT, "bot_config.json")

EMOJI_ADD    = "✅"
EMOJI_REJECT = "❌"
MAX_CARDS_PER_SCAN = 25  # maks kart na jeden skan
# ─────────────────────────────────────────────


# ── Config helpers ───────────────────────────
# Keyword entry format:
#   {"term": "shox", "max_buy_eur": 85.0, "min_discount_pct": null}
# At least one of max_buy_eur / min_discount_pct must be set.

def load_config() -> dict:
    defaults = {
        "keywords": [{"term": "shox", "max_buy_eur": 85.0, "min_discount_pct": None}],
        "max_cards": 25,
    }
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            # Migrate old single-string keyword list
            if "keywords" in data and data["keywords"] and isinstance(data["keywords"][0], str):
                old_eur = data.get("default_max_buy_eur", data.get("max_buy_eur", 85.0))
                data["keywords"] = [
                    {"term": k, "max_buy_eur": old_eur, "min_discount_pct": None}
                    for k in data["keywords"]
                ]
            # Drop obsolete global default
            data.pop("default_max_buy_eur", None)
            data.pop("max_buy_eur", None)
            return {**defaults, **data}
        except Exception:
            pass
    return defaults


def save_config(cfg: dict):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def parse_add_args(raw: str) -> dict:
    """
    Parse '!add shox eur:80 disc:40' → {"term": "shox", "max_buy_eur": 80, "min_discount_pct": 40}
    Returns None filters when none specified (caller must validate).
    """
    eur_match  = re.search(r'\beur:(\d+(?:[.,]\d+)?)\b', raw, re.IGNORECASE)
    disc_match = re.search(r'\bdisc:(\d+(?:[.,]\d+)?)\b', raw, re.IGNORECASE)

    max_buy_eur      = float(eur_match.group(1).replace(',', '.'))  if eur_match  else None
    min_discount_pct = float(disc_match.group(1).replace(',', '.')) if disc_match else None

    term = re.sub(r'\b(eur|disc):\S+', '', raw, flags=re.IGNORECASE).strip()

    return {"term": term, "max_buy_eur": max_buy_eur, "min_discount_pct": min_discount_pct}


def kw_label(kw: dict) -> str:
    """Human-readable filter description for a keyword entry."""
    parts = []
    if kw.get("max_buy_eur") is not None:
        parts.append(f"EUR≤{kw['max_buy_eur']:.0f}")
    if kw.get("min_discount_pct") is not None:
        parts.append(f"rabat≥{kw['min_discount_pct']:.0f}%")
    return " | ".join(parts) if parts else "brak filtrów"


# ── XLSX helper ──────────────────────────────

def add_to_xlsx(url: str, max_price: float) -> bool:
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_FILE)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[0].split("?")[0] == url.split("?")[0]:
            return False
    ws.append([url, max_price])
    wb.save(XLSX_FILE)
    return True


# ── Scraper helpers ──────────────────────────

def _scrape_product_sync(url: str):
    from scraper.browser import create_browser, accept_cookies
    from scraper.extractor import extract_product_data
    driver = create_browser(headless=True)
    try:
        driver.get("https://www.zalando.pl")
        time.sleep(2)
        accept_cookies(driver)
        time.sleep(1)
        return extract_product_data(driver, url)
    except Exception as e:
        logger.error("Błąd scrapowania %s: %s", url, e)
        return None
    finally:
        driver.quit()


def _should_alert(product: dict, kw_cfg: dict, eur_rate: float) -> tuple:
    """
    Returns (should_alert: bool, reason: str).
    Triggers when ANY enabled condition is met.
    """
    from scraper.pricing import calculate_buy_price_eur
    reasons = []

    max_eur = kw_cfg.get("max_buy_eur")
    if max_eur is not None:
        buy_eur = calculate_buy_price_eur(product["current_price"], eur_rate)
        product["buy_eur"] = buy_eur
        if buy_eur <= max_eur:
            reasons.append(f"buy {buy_eur:.2f} EUR ≤ {max_eur:.0f}")

    min_disc = kw_cfg.get("min_discount_pct")
    if min_disc is not None:
        disc = product.get("discount_pct") or 0
        if disc >= min_disc:
            reasons.append(f"rabat {disc}% ≥ {min_disc:.0f}%")

    if "buy_eur" not in product:
        from scraper.pricing import calculate_buy_price_eur
        product["buy_eur"] = calculate_buy_price_eur(product["current_price"], eur_rate)

    return bool(reasons), " & ".join(reasons) if reasons else ""


def _run_scan_sync(kw_cfg: dict) -> dict:
    from scraper.browser import create_browser, accept_cookies
    from scraper.pricing import get_eur_rate
    from monitor_shox import search_zalando, load_watched_urls

    term     = kw_cfg["term"]
    eur_rate = get_eur_rate()
    watched  = load_watched_urls()

    driver = create_browser(headless=True)
    try:
        driver.get("https://www.zalando.pl")
        time.sleep(2)
        accept_cookies(driver)
        time.sleep(2)
        products = search_zalando(driver, term)
    finally:
        driver.quit()

    alerts, too_expensive = [], 0
    for p in products:
        clean = p["url"].split("?")[0]
        if any(clean in w or w.split("?")[0] == clean for w in watched):
            continue
        p["eur_rate"] = eur_rate
        ok, reason = _should_alert(p, kw_cfg, eur_rate)
        if ok:
            p["alert_reason"] = reason
            alerts.append(p)
        else:
            too_expensive += 1

    return {
        "term": term,
        "total": len(products),
        "new": len(alerts) + too_expensive,
        "alerts": alerts,
        "too_expensive": too_expensive,
        "eur_rate": eur_rate,
    }


# ── Bot setup ────────────────────────────────

intents = discord.Intents.default()
intents.message_content = True
intents.reactions       = True

bot = commands.Bot(command_prefix="!", intents=intents, help_command=None)

pending_products: dict[int, dict] = {}
_scan_running = False


# ── Events ───────────────────────────────────

@bot.event
async def on_ready():
    logger.info("Bot zalogowany jako %s (%s)", bot.user, bot.user.id)
    ch = bot.get_channel(CHANNEL_ID)
    if ch:
        await ch.send("✅ **Zalando Bot uruchomiony!** Wpisz `!help` po listę komend.")
    if not daily_scan_task.is_running():
        daily_scan_task.start()
        logger.info("⏰ Zaplanowany skan dzienny o %02d:00 aktywny.", DAILY_SCAN_HOUR)


@bot.event
async def on_reaction_add(reaction: discord.Reaction, user: discord.User):
    if user.bot:
        return
    msg_id = reaction.message.id
    if msg_id not in pending_products:
        return

    product = pending_products.pop(msg_id)
    emoji   = str(reaction.emoji)

    if emoji == EMOJI_REJECT:
        try:
            await reaction.message.delete()
        except discord.NotFound:
            pass
        return

    if emoji == EMOJI_ADD:
        channel = reaction.message.channel
        try:
            await reaction.message.edit(
                content=f"⏳ Pobieram rozmiary dla **{product.get('name','?')[:60]}**..."
            )
        except Exception:
            pass

        proc = await channel.send("🔄 Otwieram stronę produktu...")
        loop = asyncio.get_event_loop()
        data = await loop.run_in_executor(None, _scrape_product_sync, product["url"])
        await proc.delete()

        price_for_xlsx = data.get("current_price", product.get("current_price", 999)) if data else product.get("current_price", 999)
        added = add_to_xlsx(product["url"], price_for_xlsx)
        xlsx_status = "✅ Dodano do xlsx" if added else "ℹ️ Już był w xlsx"

        sizes    = (data.get("sizes") or []) if data else []
        size_str = _fmt_sizes(sizes)

        embed = discord.Embed(
            title       = f"✅ DODANO DO OFERTY",
            description = (
                "\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\n"
                f"\ud83d\uded2  **BUY: {product.get('buy_eur', 0):.2f} EUR**\n"
                "\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\n"
                f"**{(data or product).get('name', 'Produkt')[:160]}**"
            ),
            url         = product["url"],
            color       = 0x57F287,
        )
        embed.add_field(name="💰 Cena",       value=f"**{price_for_xlsx:.2f} PLN**",  inline=True)
        disc = product.get("discount_pct")
        if disc:
            embed.add_field(name="🏷️ Rabat",   value=f"**−{disc}%**",                    inline=True)
        embed.add_field(name="📦 Rozmiary",   value=size_str,                         inline=False)
        embed.add_field(name="📋 xlsx",        value=xlsx_status,                      inline=False)
        if data and data.get("image_url"):
            embed.set_image(url=data["image_url"])
        embed.set_footer(text="Zalando Bot")

        await reaction.message.edit(content=None, embed=embed)
        await reaction.message.clear_reactions()


@bot.event
async def on_command_error(ctx, error):
    if isinstance(error, commands.CommandNotFound):
        return
    await ctx.send(f"❌ Błąd: `{error}`")


# ── Card sender ──────────────────────────────

async def _send_product_card(channel, product: dict):
    buy_eur  = product.get("buy_eur", 0)
    eur_rate = product.get("eur_rate", 0)
    disc     = product.get("discount_pct")
    price    = product.get("current_price", 0)
    name     = product.get("name", "Produkt")
    url      = product.get("url", "")
    reason   = product.get("alert_reason", "")

    # Color: neon green by default, gold for big discounts (≥40%)
    color = 0x00C853 if not (disc and disc >= 40) else 0xFFD700

    # TITLE = BUY PRICE (biggest text in embed, also the clickable link)
    # Description = product name
    embed = discord.Embed(
        title       = f"🛒  BUY:  {buy_eur:.2f} EUR",
        url         = url,
        description = f"**{name[:120]}**",
        color       = color,
    )
    embed.add_field(name="💰 Cena Zalando", value=f"**{price:.0f} PLN**", inline=True)
    if disc:
        orig = product.get("original_price")
        orig_str = f"\n~~{orig:.0f} PLN~~" if orig else ""
        embed.add_field(name="🏷️ Rabat", value=f"**−{disc}%**{orig_str}", inline=True)
    embed.add_field(name="📈 Kurs", value=f"{eur_rate:.4f} PLN/EUR", inline=True)
    if reason:
        embed.add_field(name="🔔 Trigger", value=f"`{reason}`", inline=False)
    if product.get("image_url"):
        embed.set_image(url=product["image_url"])
    embed.set_footer(text="✅  Dodaj do oferty      ❌  Odrzuć")

    msg = await channel.send(embed=embed)
    await msg.add_reaction(EMOJI_ADD)
    await msg.add_reaction(EMOJI_REJECT)
    pending_products[msg.id] = product
    return msg



# ─── Size formatter ──────────────────────────

def _fmt_sizes(sizes: list) -> str:
    """
    Format size list for Discord embed.
    Dicts: {'size': '40', 'available': True, 'low_stock': False}
    Available sizes in backticks, ⚠️ = low stock, 5 per row.
    """
    if not sizes:
        return "—"

    if not isinstance(sizes[0], dict):
        parts = [f"`{s}`" for s in sizes[:20]]
    else:
        parts = []
        for s in sizes:
            if not s.get("available"):
                continue
            label = str(s["size"])
            if s.get("low_stock"):
                label += "⚠️"
            parts.append(f"`{label}`")

    if not parts:
        return "_Brak dostępnych rozmiarów_"

    # 5 sizes per row
    rows = []
    for i in range(0, len(parts), 5):
        rows.append("  ".join(parts[i:i + 5]))
    result = "\n".join(rows)
    return result[:1000]  # Discord field value limit = 1024


# ── Background size prefetch ──────────────────────────


def _prefetch_sizes_batch_sync(products: list) -> dict:
    """
    Open ONE Chrome session, visit each product URL, extract sizes.
    Returns dict {url: [size_str, ...]}
    """
    from scraper.browser import create_browser, accept_cookies
    from scraper.extractor import extract_product_data

    driver = create_browser(headless=True)
    result = {}
    try:
        driver.get("https://www.zalando.pl")
        time.sleep(2)
        accept_cookies(driver)
        time.sleep(1)
        for product in products:
            url = product["url"]
            try:
                data = extract_product_data(driver, url)
                result[url]           = data.get("sizes", []) if data else []
                result[url + "_data"] = data  # full data: current_price, discount_pct, etc.
            except Exception as e:
                logger.warning("Size prefetch error %s: %s", url, e)
                result[url]           = []
                result[url + "_data"] = None
            time.sleep(1.2)
    except Exception as e:
        logger.error("Prefetch session error: %s", e)
    finally:
        try:
            driver.quit()
        except Exception:
            pass
    return result


async def _background_fetch_sizes(channel, card_msgs: list, products: list):
    """Background task: fetch sizes for all cards and edit embeds."""
    await asyncio.sleep(0.5)
    status_msg = await channel.send(
        f"⏳ Pobieram rozmiary dla **{len(products)}** produktów w tle... (~{len(products)*2}s)"
    )
    loop = asyncio.get_event_loop()
    try:
        sizes_map = await loop.run_in_executor(None, _prefetch_sizes_batch_sync, products)
    except Exception as e:
        logger.error("Background size fetch failed: %s", e)
        await status_msg.edit(content=f"❌ Błąd przy pobieraniu rozmiarów: `{e}`")
        return

    edited = 0
    for msg, product in zip(card_msgs, products):
        url      = product["url"]
        sizes    = sizes_map.get(url, [])
        size_str = _fmt_sizes(sizes)

        # Rebuild the card embed with sizes AND accurate price from product page
        try:
            if msg.id not in pending_products:
                continue  # product was already acted on
            old_embed = msg.embeds[0] if msg.embeds else None
            if old_embed is None:
                continue

            data        = sizes_map.get(url + "_data")   # extra: full data cached below
            eur_rate    = product.get("eur_rate", 0)
            name        = product.get("name", "Produkt")
            reason      = product.get("alert_reason", "")

            # Use real price from product page if available
            real_price  = (data.get("current_price") if data else None) or product.get("current_price", 0)
            real_orig   = (data.get("original_price") if data else None) or product.get("original_price")
            # extractor uses "discount_percent", listing uses "discount_pct" — handle both
            real_disc   = (data.get("discount_percent") or data.get("discount_pct") if data else None) or product.get("discount_pct")
            real_img    = (data.get("image_url")       if data else None) or product.get("image_url")
            real_name   = (data.get("name")            if data else None) or name

            # Recalculate buy_eur if price changed
            if eur_rate and real_price:
                from scraper.pricing import calculate_buy_price_eur
                buy_eur = calculate_buy_price_eur(real_price, eur_rate)
                # Update cached product so ✅ handler uses correct price
                pending_products[msg.id]["current_price"] = real_price
                pending_products[msg.id]["buy_eur"]       = buy_eur
            else:
                buy_eur = product.get("buy_eur", 0)

            color = 0x00C853 if not (real_disc and real_disc >= 40) else 0xFFD700

            embed = discord.Embed(
                title       = f"🛒  BUY:  {buy_eur:.2f} EUR",
                url         = url,
                description = f"**{real_name[:120]}**",
                color       = color,
            )
            embed.add_field(name="💰 Cena Zalando", value=f"**{real_price:.0f} PLN**", inline=True)
            if real_disc:
                orig_str = f"\n~~{real_orig:.0f} PLN~~" if real_orig else ""
                embed.add_field(name="🏷️ Rabat", value=f"**−{real_disc}%**{orig_str}", inline=True)
            embed.add_field(name="📈 Kurs", value=f"{eur_rate:.4f} PLN/EUR", inline=True)
            embed.add_field(name="📦 Rozmiary", value=size_str, inline=False)
            if reason:
                embed.add_field(name="🔔 Trigger", value=f"`{reason}`", inline=False)
            if real_img:
                embed.set_image(url=real_img)
            embed.set_footer(text="✅  Dodaj do oferty      ❌  Odrzuć")

            await msg.edit(embed=embed)
            edited += 1
            await asyncio.sleep(0.5)
        except discord.NotFound:
            pass  # message was deleted (user rejected it)
        except Exception as e:
            logger.warning("Embed edit error for %s: %s", url, e)

    if edited:
        await status_msg.edit(content=f"📦 Rozmiary zaktualizowane — **{edited}** kart. (⚠️ = mało sztuk)")
    else:
        await status_msg.edit(content="📦 Rozmiary pobrane, brak kart do zaktualizowania.")


# ── Scan runner ──────────────────────────────

async def run_scan_async(channel, kw_cfg: dict):
    global _scan_running
    if _scan_running:
        await channel.send("⏳ Skan już trwa.")
        return

    _scan_running = True
    filters = kw_label(kw_cfg)
    status  = await channel.send(f"🔍 Skanuję **{kw_cfg['term']}** ({filters})... chwilę.")

    try:
        results = await asyncio.get_event_loop().run_in_executor(None, _run_scan_sync, kw_cfg)
    except Exception as e:
        await status.edit(content=f"❌ Błąd: `{e}`")
        logger.error("Skan error: %s", e, exc_info=True)
        _scan_running = False
        return

    _scan_running = False
    alerts = results["alerts"]

    await status.edit(content=(
        f"✅ **{results['term']}** — "
        f"znaleziono {results['total']} | nowe {results['new']} | "
        f"alerty {len(alerts)} | za drogie {results['too_expensive']}"
    ))

    if not alerts:
        await channel.send("🔕 Brak produktów spełniających kryteria.")
        return

    await channel.send(f"👇 Kliknij **✅** żeby dodać do oferty, **❌** żeby odrzucić:")
    max_cards = load_config().get("max_cards", 25)
    shown     = alerts[:max_cards]
    card_msgs = []
    for p in shown:
        msg = await _send_product_card(channel, p)
        card_msgs.append(msg)
        await asyncio.sleep(1.0)
    if len(alerts) > max_cards:
        await channel.send(f"_...i jeszcze {len(alerts) - max_cards} produktów (pokazano {max_cards}). Zmień limit: `!limit <liczba>`_")

    # Start background size prefetch
    asyncio.create_task(_background_fetch_sizes(channel, card_msgs, shown))


# ── Commands ───────────────────────────────

@bot.command(name="help")
async def cmd_help(ctx):
    embed = discord.Embed(title="📋 Zalando Bot — komendy", color=0x5865F2)
    embed.add_field(name="!scan [keyword]",          value="Skanuj keyword (lub pierwszy z listy)", inline=False)
    embed.add_field(name="!scanall",                 value="Skanuj wszystkie keywordy", inline=False)
    embed.add_field(name="!keywords",                value="Pokaż listę keywordów z filtrami", inline=False)
    embed.add_field(name="!add <kw> [eur:X] [disc:X]", value=(
        "Dodaj keyword\n"
        "`!add shox` — domyślny próg EUR\n"
        "`!add shox eur:80` — alert gdy buy ≤ 80 EUR\n"
        "`!add shox disc:40` — alert gdy rabat ≥ 40%\n"
        "`!add jordan eur:90 disc:35` — jedno z dwóch"
    ), inline=False)
    embed.add_field(name="!remove <keyword>",        value="Usuń keyword z listy", inline=False)
    embed.add_field(name="!prog <kwota>",            value="Zmień domyślny próg EUR", inline=False)
    embed.add_field(name="!status",                  value="Aktualny config", inline=False)
    embed.add_field(name="Reakcje:",                 value="✅ Dodaj do oferty  |  ❌ Odrzuć", inline=False)
    await ctx.send(embed=embed)


@bot.command(name="status")
async def cmd_status(ctx):
    cfg = load_config()
    kws = "\n".join(f"• `{k['term']}` — {kw_label(k)}" for k in cfg["keywords"]) or "_(brak)_"
    embed = discord.Embed(title="⚙️ Config", color=0xFEE75C)
    embed.add_field(name="🔎 Keywordy",        value=kws,                                              inline=False)
    embed.add_field(name="💶 Domyślny próg",   value=f"`{cfg['default_max_buy_eur']} EUR`",            inline=True)
    embed.add_field(name="⏳ Skan trwa",        value="✅" if _scan_running else "❌",                 inline=True)
    embed.add_field(name="⏸ Oczekuje",         value=f"{len(pending_products)} kart",                 inline=True)
    await ctx.send(embed=embed)


@bot.command(name="keywords")
async def cmd_keywords(ctx):
    cfg = load_config()
    if not cfg["keywords"]:
        await ctx.send("📋 Brak keywordów. Dodaj: `!add shox`")
        return
    lines = "\n".join(
        f"{i+1}. **{k['term']}** — {kw_label(k)}"
        for i, k in enumerate(cfg["keywords"])
    )
    await ctx.send(f"📋 **Keywordy ({len(cfg['keywords'])}):**\n{lines}")


@bot.command(name="add")
async def cmd_add(ctx, *, args: str = None):
    if not args:
        await ctx.send("❌ Użycie: `!add <keyword> [eur:X] [disc:X]`\nPrzykłady:\n`!add shox eur:85`\n`!add air max disc:40`")
        return
    cfg    = load_config()
    parsed = parse_add_args(args)
    term   = parsed["term"].lower()
    if not term:
        await ctx.send("❌ Podaj nazwę keywordu.")
        return
    if parsed["max_buy_eur"] is None and parsed["min_discount_pct"] is None:
        await ctx.send(
            "❌ Podaj filtr:\n"
            "`!add shox eur:85` — buy price ≤ 85 EUR\n"
            "`!add shox disc:40` — rabat ≥ 40%\n"
            "`!add jordan eur:90 disc:35` — jedno z dwóch"
        )
        return
    if any(k["term"].lower() == term for k in cfg["keywords"]):
        await ctx.send(f"ℹ️ `{term}` już jest na liście.")
        return
    cfg["keywords"].append({"term": term, "max_buy_eur": parsed["max_buy_eur"], "min_discount_pct": parsed["min_discount_pct"]})
    save_config(cfg)
    filters = kw_label(parsed)
    await ctx.send(f"✅ Dodano `{term}` [{filters}]. Razem: {len(cfg['keywords'])} keywordów.")


@bot.command(name="edit")
async def cmd_edit(ctx, *, args: str = None):
    """Zmień filtry istniejącego keywordu. Użycie: !edit shox eur:80"""
    if not args:
        await ctx.send(
            "❌ Użycie: `!edit <keyword> [eur:X] [disc:X]`\n"
            "Przykłady:\n"
            "`!edit shox eur:80`\n"
            "`!edit shox disc:45`\n"
            "`!edit jordan eur:90 disc:30`"
        )
        return
    cfg    = load_config()
    parsed = parse_add_args(args)
    term   = parsed["term"].lower()

    if not term:
        await ctx.send("❌ Podaj nazwę keywordu.")
        return
    if parsed["max_buy_eur"] is None and parsed["min_discount_pct"] is None:
        await ctx.send("❌ Podaj nowy filtr: `eur:X` i/lub `disc:X`")
        return

    idx = next((i for i, k in enumerate(cfg["keywords"]) if k["term"].lower() == term), None)
    if idx is None:
        await ctx.send(f"❌ Keyword `{term}` nie istnieje. Dodaj go: `!add {term} eur:85`")
        return

    old_label = kw_label(cfg["keywords"][idx])
    if parsed["max_buy_eur"] is not None:
        cfg["keywords"][idx]["max_buy_eur"] = parsed["max_buy_eur"]
    if parsed["min_discount_pct"] is not None:
        cfg["keywords"][idx]["min_discount_pct"] = parsed["min_discount_pct"]
    new_label = kw_label(cfg["keywords"][idx])
    save_config(cfg)
    await ctx.send(f"✅ `{term}` zaktualizowany: `{old_label}` → `{new_label}`")


@bot.command(name="remove")
async def cmd_remove(ctx, *, keyword: str = None):
    if not keyword:
        await ctx.send("❌ Użycie: `!remove <keyword>`")
        return
    keyword = keyword.strip().lower()
    cfg = load_config()
    before = len(cfg["keywords"])
    cfg["keywords"] = [k for k in cfg["keywords"] if k["term"].lower() != keyword]
    if len(cfg["keywords"]) == before:
        await ctx.send(f"❌ `{keyword}` nie ma na liście.")
        return
    save_config(cfg)
    await ctx.send(f"✅ Usunięto `{keyword}`. Zostało: {len(cfg['keywords'])}.")


@bot.command(name="prog")
async def cmd_prog(ctx, *args):
    await ctx.send(
        "❌ `!prog` jest już nieaktywny. Progi ustawiaj per-keyword:\n"
        "`!add shox eur:85` lub `!add shox disc:40`\n"
        "Jeśli chcesz zmienić istniejący keyword: `!remove shox`, potem `!add shox eur:80`"
    )


@bot.command(name="limit")
async def cmd_limit(ctx, amount: str = None):
    if not amount:
        cfg = load_config()
        await ctx.send(f"ℹ️ Aktualny limit kart per skan: **{cfg.get('max_cards', 10)}**. Zmień: `!limit <liczba>`")
        return
    try:
        value = int(amount)
        if value < 1 or value > 50:
            raise ValueError
    except ValueError:
        await ctx.send("❌ Podaj liczbę od 1 do 50.")
        return
    cfg = load_config()
    cfg["max_cards"] = value
    save_config(cfg)
    await ctx.send(f"✅ Limit kart per skan: **{value}**")


@bot.command(name="scan")
async def cmd_scan(ctx, *, keyword: str = None):
    cfg = load_config()
    if keyword:
        kw_str = keyword.strip()
        if kw_str.startswith("http"):
            await ctx.send(
                "❌ `!scan` przyjmuje słowo kluczowe (np. `!scan shox`), nie link.\n"
                "Jeśli chcesz dodać konkretny but, kliknij **✅** na jego karcie po skanie."
            )
            return
        parsed = parse_add_args(kw_str)
        # If no filters given, look up saved keyword
        if parsed["max_buy_eur"] is None and parsed["min_discount_pct"] is None:
            term_lower = parsed["term"].lower()
            saved = next((k for k in cfg["keywords"] if k["term"].lower() == term_lower), None)
            if saved:
                kw_cfg = saved
            else:
                await ctx.send(
                    f"❌ Keyword `{parsed['term']}` nie ma zapisanego progu.\n"
                    f"Dodaj go najpierw: `!add {parsed['term']} eur:85` lub `!add {parsed['term']} disc:40`"
                )
                return
        else:
            kw_cfg = parsed
    elif cfg["keywords"]:
        kw_cfg = cfg["keywords"][0]
    else:
        await ctx.send("❌ Brak keywordów. Dodaj: `!add shox eur:85`")
        return
    await run_scan_async(ctx.channel, kw_cfg)


@bot.command(name="scanall")
async def cmd_scanall(ctx):
    cfg = load_config()
    if not cfg["keywords"]:
        await ctx.send("❌ Brak keywordów. Dodaj: `!add shox`")
        return
    if _scan_running:
        await ctx.send("⏳ Skan już trwa.")
        return
    names = ", ".join(f"`{k['term']}`" for k in cfg["keywords"])
    await ctx.send(f"🔍 Skanuję {len(cfg['keywords'])} keywordów: {names}")
    for kw in cfg["keywords"]:
        await run_scan_async(ctx.channel, kw)
        await asyncio.sleep(2)


@bot.command(name="importlinks")
async def cmd_import_links(ctx):
    """
    Attach .xlsx in the same format as produkty.xlsx (URL | Max Cena).
    Bot adds all new URLs to produkty.xlsx.
    """
    if not ctx.message.attachments:
        await ctx.send("❌ Załącz plik .xlsx z URL-ami. Format: dwie kolumny — URL | Max Cena.")
        return
    att = ctx.message.attachments[0]
    if not att.filename.lower().endswith(".xlsx"):
        await ctx.send("❌ Plik musi być w formacie .xlsx")
        return

    import openpyxl
    raw = await att.read()
    src_wb = openpyxl.load_workbook(io.BytesIO(raw))
    src_ws = src_wb.active

    dst_wb = openpyxl.load_workbook(XLSX_FILE)
    dst_ws = dst_wb.active

    # Existing URLs for dedup
    existing = set()
    for row in dst_ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            existing.add(str(row[0]).split("?")[0].strip())

    added = skipped = errors = 0
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        url = str(row[0]).strip()
        if not url.startswith("http"):
            errors += 1
            continue
        max_price = float(row[1]) if len(row) > 1 and row[1] is not None else 999.0
        clean = url.split("?")[0]
        if clean in existing:
            skipped += 1
        else:
            dst_ws.append([url, max_price])
            existing.add(clean)
            added += 1

    dst_wb.save(XLSX_FILE)
    await ctx.send(
        f"✅ Import zakończony z `{att.filename}`:\n"
        f"• Dodano: **{added}** URL-i\n"
        f"• Pominięto (duplikaty): {skipped}\n"
        f"• Błędy (nieprawidłowe URL): {errors}"
    )


@bot.command(name="removelinks")
async def cmd_remove_links(ctx):
    """
    Attach .xlsx with URLs to remove from produkty.xlsx.
    Only column A (URL) is checked — Max Cena is ignored.
    """
    if not ctx.message.attachments:
        await ctx.send("❌ Załącz plik .xlsx z URL-ami do usunięcia.")
        return
    att = ctx.message.attachments[0]
    if not att.filename.lower().endswith(".xlsx"):
        await ctx.send("❌ Plik musi być w formacie .xlsx")
        return

    import openpyxl
    raw = await att.read()
    src_wb = openpyxl.load_workbook(io.BytesIO(raw))
    src_ws = src_wb.active

    # URLs to remove
    to_remove = set()
    for row in src_ws.iter_rows(min_row=1, values_only=True):
        if row and row[0] and str(row[0]).startswith("http"):
            to_remove.add(str(row[0]).split("?")[0].strip())

    if not to_remove:
        await ctx.send("❌ Nie znaleziono żadnych URL-i w pliku.")
        return

    dst_wb = openpyxl.load_workbook(XLSX_FILE)
    dst_ws = dst_wb.active

    # Rebuild sheet without removed URLs
    rows_to_keep = []
    for row in dst_ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        url = str(row[0]).split("?")[0].strip() if row[0] else ""
        if url in to_remove:
            continue
        rows_to_keep.append(list(row))

    removed = dst_ws.max_row - len(rows_to_keep)

    # Clear sheet and rewrite
    dst_ws.delete_rows(1, dst_ws.max_row)
    for r in rows_to_keep:
        dst_ws.append(r)

    dst_wb.save(XLSX_FILE)
    await ctx.send(
        f"✅ Usunięto **{removed}** URL-i z `produkty.xlsx`."
    )


@bot.command(name="listlinks")
async def cmd_list_links(ctx, page: int = 1):
    """Show URLs currently in produkty.xlsx."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True)
    ws = wb.active
    rows = [(row[0], row[1]) for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]]
    wb.close()

    per_page = 15
    total_pages = max(1, (len(rows) + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    chunk = rows[start:start + per_page]

    lines = []
    for i, (url, price) in enumerate(chunk, start=start + 1):
        short = (url or "").split("/")[-1].replace(".html", "")[:50]
        lines.append(f"{i}. `{short}` — {price} PLN")

    embed = discord.Embed(
        title=f"📊 produkty.xlsx ({len(rows)} URL-i) — strona {page}/{total_pages}",
        description="\n".join(lines) or "_(puste)_",
        color=0x5865F2,
    )
    embed.set_footer(text=f"Następna strona: `!listlinks {page + 1}`" if page < total_pages else "Ostatnia strona")
    await ctx.send(embed=embed)


@bot.command(name="stats")
async def cmd_stats(ctx):
    """Statystyki produkty.xlsx — podział na keywordy, zakresy cen."""
    import openpyxl
    cfg = load_config()
    wb = openpyxl.load_workbook(XLSX_FILE, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] and str(row[0]).startswith("http"):
            rows.append((str(row[0]).strip(), float(row[1]) if len(row) > 1 and row[1] else None))
    wb.close()

    total = len(rows)
    keywords = [k["term"] for k in cfg["keywords"]]

    # Group by first matching keyword
    groups = {}
    for url, price in rows:
        cat = next((kw for kw in keywords if kw.lower() in url.lower()), "Inne")
        groups[cat] = groups.get(cat, 0) + 1

    # Price range buckets (max_price column)
    buckets = {"≤200 PLN": 0, "201-500 PLN": 0, "501-1000 PLN": 0, ">1000 PLN": 0}
    prices = []
    for _, price in rows:
        if price is None:
            continue
        prices.append(price)
        if price <= 200:      buckets["≤200 PLN"] += 1
        elif price <= 500:    buckets["201-500 PLN"] += 1
        elif price <= 1000:   buckets["501-1000 PLN"] += 1
        else:                 buckets[">1000 PLN"] += 1

    embed = discord.Embed(title="📊 Statystyki produkty.xlsx", color=0x57F287)
    embed.add_field(name="📦 Łącznie URL-i", value=f"**{total}**", inline=True)
    avg = sum(prices) / len(prices) if prices else 0
    embed.add_field(name="💰 Śr. max cena", value=f"**{avg:.0f} PLN**" if avg else "_brak_", inline=True)
    embed.add_field(name="​", value="​", inline=True)  # spacer

    # Keyword breakdown
    kw_lines = sorted(groups.items(), key=lambda x: -x[1])
    kw_text = "\n".join(
        f"🔹 **{cat}**: {cnt} ({cnt/total*100:.0f}%)" for cat, cnt in kw_lines if cnt > 0
    ) or "_(brak)_"
    embed.add_field(name="🔎 Podział wg keyword", value=kw_text, inline=False)

    # Price breakdown
    bucket_text = "\n".join(f"**{k}**: {v}" for k, v in buckets.items() if v > 0)
    if bucket_text:
        embed.add_field(name="💶 Zakresy max ceny", value=bucket_text, inline=False)

    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(XLSX_FILE)).strftime("%d.%m.%Y %H:%M")
    embed.set_footer(text=f"produkty.xlsx · ostatnia zmiana: {mtime}")
    await ctx.send(embed=embed)


@bot.command(name="raport")
async def cmd_raport(ctx):
    """Generuje i wysyła raport.xlsx z podziałem na keywordy."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    cfg = load_config()

    wb_src = openpyxl.load_workbook(XLSX_FILE, read_only=True)
    ws_src = wb_src.active
    all_rows = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row and row[0] and str(row[0]).startswith("http"):
            url   = str(row[0]).strip()
            price = float(row[1]) if len(row) > 1 and row[1] else None
            slug  = url.split("/")[-1].replace(".html", "")
            all_rows.append({"url": url, "price": price, "slug": slug})
    wb_src.close()

    keywords = [k["term"] for k in cfg["keywords"]]

    def get_cat(url):
        return next((kw for kw in keywords if kw.lower() in url.lower()), "Inne")

    for r in all_rows:
        r["cat"] = get_cat(r["url"])

    by_cat = {}
    for r in all_rows:
        by_cat.setdefault(r["cat"], []).append(r)

    # ── Build report workbook ──
    wb  = openpyxl.Workbook()
    h_f = Font(bold=True, color="FFFFFF")
    h_b = PatternFill("solid", fgColor="1a6b3a")
    sh_b = PatternFill("solid", fgColor="C6EFCE")

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "Podsumowanie"
    ws["A1"] = "Raport Zalando Monitor"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = f"Wygenerowany: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A3"] = f"Łącznie produktów: {len(all_rows)}"
    ws["A3"].font = Font(bold=True)

    for col, label in zip("ABC", ["Keyword", "Liczba", "% całości"]):
        c = ws[f"{col}5"]
        c.value = label
        c.font  = h_f
        c.fill  = h_b

    for i, (cat, items) in enumerate(sorted(by_cat.items(), key=lambda x: -len(x[1])), start=6):
        pct = len(items) / len(all_rows) * 100 if all_rows else 0
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=len(items))
        ws.cell(row=i, column=3, value=f"{pct:.1f}%")

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12

    # Sheet per keyword
    for cat in list(keywords) + ["Inne"]:
        items = by_cat.get(cat, [])
        if not items:
            continue
        wk = wb.create_sheet(title=f"{cat[:25]} ({len(items)})")
        for col, label in zip("ABC", ["URL", "Max Cena (PLN)", "Slug"]):
            c = wk[f"{col}1"]
            c.value = label
            c.font  = Font(bold=True)
            c.fill  = sh_b
        for i, r in enumerate(items, start=2):
            wk.cell(row=i, column=1, value=r["url"])
            wk.cell(row=i, column=2, value=r["price"])
            wk.cell(row=i, column=3, value=r["slug"])
        wk.column_dimensions["A"].width = 80
        wk.column_dimensions["B"].width = 15
        wk.column_dimensions["C"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"raport_zalando_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    summary = " | ".join(f"{cat}: {len(items)}" for cat, items in sorted(by_cat.items(), key=lambda x: -len(x[1])))
    await ctx.send(
        content=f"📊 Raport — **{len(all_rows)}** produktów\n{summary}",
        file=discord.File(buf, filename=fname),
    )


# ── Scheduled daily scan ─────────────────────

# Runs at 01:00 Warsaw time (UTC+1 in winter, UTC+2 in summer)
# Using UTC+1 for simplicity — adjust hour if needed

DAILY_SCAN_HOUR = 1  # 01:00

@tasks.loop(time=datetime.time(hour=DAILY_SCAN_HOUR, minute=0,
                               tzinfo=datetime.timezone(datetime.timedelta(hours=1))))
async def daily_scan_task():
    channel = bot.get_channel(CHANNEL_ID)
    if not channel:
        return
    cfg = load_config()
    if not cfg["keywords"]:
        return

    now = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    await channel.send(f"⏰ **Automatyczny skan dzienny — {now}**")

    total_alerts = 0
    all_results = []
    for kw in cfg["keywords"]:
        await run_scan_async(channel, kw)
        await asyncio.sleep(3)



# ── Start ─────────────────────────────────────

if __name__ == "__main__":
    bot.run(BOT_TOKEN)
