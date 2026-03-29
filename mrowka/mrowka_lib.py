from __future__ import annotations
from typing import Optional, Union
import aiofiles  # type: ignore
import dc
import mrowka_data
import common
import aiohttp
import discord
import mails
import logger
import gmail_imap
import asyncio
import json
import ean_db
# import eans.ean
from discord.ext import commands

import sys
import os
import json as _json

# Ścieżka do Sfera CLI
_SFERA_EXE = os.path.join(os.path.dirname(__file__), "..", "ReflectSfera", "bin", "Release", "net8.0-windows", "ReflectSfera.exe")
_SFERA_DB_NAME = "Nexo_eleat teesty kurwa"
_SFERA_DB_SERVER = r".\INSERTNEXO"
_SFERA_PASSWORD = ""


async def _sfera_cli(action: str, data: dict, timeout: int = 120) -> dict:
    """Wywołuje ReflectSfera.exe asynchronicznie. Zwraca dict odpowiedzi."""
    req = {
        "Action": action,
        "DbName": _SFERA_DB_NAME,
        "DbServer": _SFERA_DB_SERVER,
        "SferaPassword": _SFERA_PASSWORD,
        **data
    }
    import tempfile, pathlib
    tmp = pathlib.Path(tempfile.mktemp(suffix=".json", dir=os.path.dirname(__file__)))
    try:
        tmp.write_text(_json.dumps(req, ensure_ascii=False), encoding="utf-8")
        proc = await asyncio.create_subprocess_exec(
            _SFERA_EXE, str(tmp),
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        try:
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        except asyncio.TimeoutError:
            proc.kill()
            return {"Success": False, "Message": f"Timeout ({timeout}s) w ReflectSfera"}
        out = stdout.decode("utf-8", errors="replace").strip()
        for line in reversed(out.splitlines()):
            line = line.strip()
            if line.startswith("{") and "Success" in line:
                try:
                    return _json.loads(line)
                except Exception:
                    pass
        logger.logger.error("_sfera_cli [%s] błędna odpowiedź: %s", action, out[:300])
        return {"Success": False, "Message": "Zła odpowiedź CLI", "Raw": out}
    except Exception as e:
        logger.logger.error("_sfera_cli [%s] wyjątek: %s", action, e)
        return {"Success": False, "Message": str(e)}
    finally:
        try: tmp.unlink()
        except: pass


# --- Stara funkcja zachowana dla kompatybilności z mrowka_bot.py ---
async def _subiekt_post(endpoint: str, payload: dict, timeout: int = 30) -> Optional[dict]:
    """Mapuje stare wywołania HTTP na lokalny Sfera CLI."""
    ACTION_MAP = {
        "/api/products/ensure": ("EnsureProducts", lambda p: {"EnsureProducts": [p]}),
        "/api/pz/create":       ("CreatePZ",       lambda p: {"PzData": p}),
        "/api/pz/update-tracking": ("UpdateTracking", lambda p: {"UpdateTracking": p}),
        "/api/pz/update-invoice":  ("UpdateInvoice",  lambda p: {"UpdateInvoice": p}),
        "/api/pz/accept":           ("AcceptPZ",        lambda p: {"AcceptPz": p}),
        "/api/fz/create":           ("CreateFZ",        lambda p: {"FzData": p}),
        "/api/pz/update-uwagi":     ("UpdatePzUwagi",   lambda p: {"UpdatePzUwagi": p}),
    }
    if endpoint not in ACTION_MAP:
        logger.logger.warning("_subiekt_post: nieznany endpoint %s", endpoint)
        return None
    action, mapper = ACTION_MAP[endpoint]
    res = await _sfera_cli(action, mapper(payload), timeout=timeout)
    if res.get("Success"):
        # Mapuj odpowiedź CLI na format oczekiwany przez stary kod
        if action in ("CreatePZ", "CreateFZ"):
            return {"Success": True, "documentNumber": res.get("DocumentNumber", "?")}
        if action == "AcceptPZ":
            return {"sygnatura": res.get("Sygnatura", "?")}
        if action == "EnsureProducts":
            return {"created": True, "message": "OK"}
        return res
    else:
        logger.logger.error("_subiekt_post [%s]: %s", endpoint, res.get("Message", "?"))
        return None


async def ensure_products_in_subiekt(ean_entries: list["ean_db.EanEntry"]) -> dict[str, bool]:
    """
    Sprawdza/tworzy produkty w Subiekcie przez lokalne Sfera CLI.

    FLOW DODAWANIA PRODUKTÓW DO SUBIEKT:
    =====================================
    1. zalando_scanner.py skanuje URL przez Selenium i zbiera EAN dla WSZYSTKICH dostępnych
       rozmiarów (InStock) ze strony Zalando (używa JSON-LD scraping w extractor.py).
    2. Wyniki są zapisywane w lokalnym SQLite (ean_db.sqlite) — jeden wiersz per link+rozmiar.
    3. Przy tworzeniu PZ, ean_db.get_eans_for_links() zwraca WSZYSTKIE rozmiary dla danego
       linku produktu (nie tylko te z zamówienia).
    4. Ta funkcja (ensure_products_in_subiekt) wysyła WSZYSTKIE te rozmiary do EnsureProducts
       w ReflectSfera.exe — dzięki czemu przy NOWYM produkcie tworzone są warianty rozmiaru
       dla wszystkich dostępnych rozmiarów w Zalando, nie tylko dla zamówionego.
    5. ReflectSfera pomija rozmiary które już istnieją w Subiekcie (idempotentne).

    PRZYKŁAD: Zamówiono Nike Air Max 90 rozmiar 43.
    - ean_db ma wpisy dla rozmiarów 41, 42, 42.5, 43, 44... (wszystkie InStock z Zalando)
    - EnsureProducts tworzy w Subiekcie produkt z WSZYSTKIMI tymi wariantami
    - PZ tworzone tylko dla pozycji z zamówienia (rozmiar 43, ilość X)

    Args:
        ean_entries: lista EanEntry — WSZYSTKIE rozmiary z ean_db dla danego produktu/linku

    Returns:
        dict {ean_or_key: bool} wskazujący sukces per wpis
    """
    results: dict[str, bool] = {}
    sfera_entries = [
        {
            "EAN": e.ean,
            "SKU": e.sku,
            "Brand": e.brand,
            # Format nazwy w Subiekcie: "Nike Sportswear HV4517-100.43"
            "ModelName": f"{e.brand} {e.sku} \u00b7 {e.size}" if (e.brand and e.sku) else (e.name or f"{e.brand} \u00b7 {e.size}"),
            "Size": e.size,
        }
        for e in ean_entries if (e.ean or e.sku)
    ]
    if not sfera_entries:
        return results
    res = await _sfera_cli("EnsureProducts", {"EnsureProducts": sfera_entries}, timeout=180)
    if res.get("Success") and res.get("EnsureResults"):
        for k, v in res["EnsureResults"].items():
            results[k] = v
    else:
        for e in ean_entries:
            key = f"{e.ean or e.name or '?'}:{e.size}"
            results[key] = False
        logger.logger.warning("ensure_products_in_subiekt CLI error: %s", res.get("Message"))
    return results



async def _create_pz_for_order_item(
    order_item: "mrowka_data.MrowkaOrderItem",
    email: Optional[str],
) -> Optional[str]:
    """
    Tworzy PZ (Przyjęcie Zewnętrzne) w Subiekcie dla danego podzielonego zamówienia.

    Flow:
    1. Pobierz EANy z ean_db dla linków w order_item.shoes
    2. Dla każdego EAN sprawdź/utwórz produkt w Subiekcie (ensure_products_in_subiekt)
    3. Stwórz PZ z listą pozycji {ean, quantity, cenaZakupu}
    4. Zwróć numer dokumentu PZ

    Returns: numer dokumentu PZ lub None przy błędzie.
    """
    # 1. Zbierz wszystkie linki i pobierz ich EANy z lokalnej bazy
    links = [shoe.link for shoe in order_item.shoes]
    ean_map = await ean_db.get_eans_for_links(links)  # {link: [EanEntry]}

    # 2. Zbuduj listę pozycji PZ — jeden wpis per (EAN lub nazwa, rozmiar, ilość)
    pz_items: list[dict] = []
    all_entries: list[ean_db.EanEntry] = []

    for shoe in order_item.shoes:
        entries = ean_map.get(shoe.link, [])
        all_entries.extend(entries)
        # size_to_quantity z shoe informuje o ilościach dla różnych rozmiarów
        for size, quantity in shoe.size_to_quantity.items():
            if quantity <= 0:
                continue
            # Znajdź entry dla tego rozmiaru
            entry = next((e for e in entries if e.size == size), None)
            pz_items.append({
                "ean": entry.ean if entry else None,
                "symbol": entry.sku if entry else None,
                # Nazwa jako fallback gdy brak EAN (backend szuka po nazwie)
                "nazwaFallback": f"{entry.brand or ''} {entry.name or ''} • {size}".strip() if entry else size,
                "quantity": quantity,
                "cenaZakupu": entry.price if entry else shoe.price,
            })

    if not pz_items:
        logger.logger.warning("_create_pz_for_order_item: brak pozycji dla %s", order_item.name)
        return None

    # 3. Sprawdź/utwórz produkty w Subiekcie (deduplikacja po EAN lub nazwie+rozmiar)
    if all_entries:
        await ensure_products_in_subiekt(all_entries)

    # 4. Utwórz PZ
    uwagi = f"Zamówienie Zalando: {order_item.name}"
    if email:
        uwagi += f" | Konto: {email}"

    payload = {
        "dostawcaNip": None,
        "dostawcaEmail": None,
        "uwagi": uwagi,
        "items": [
            {
                "ean": item["ean"],
                "symbol": item["symbol"],
                "quantity": item["quantity"],
                "cenaZakupu": item["cenaZakupu"],
            }
            for item in pz_items
        ],
    }

    result = await _sfera_cli("CreatePZ", {"PzData": payload}, timeout=120)
    if result.get("Success") and result.get("DocumentNumber"):
        doc_num = result["DocumentNumber"]
        logger.logger.info("✅ PZ %s → zamowienie %s", doc_num, order_item.name)
        return doc_num
    else:
        logger.logger.error("❌ Błąd tworzenia PZ dla: %s | result=%s", order_item.name, result)
        return None


async def save_attachment_to_file(attachment: discord.Attachment) -> str:
    file_path = common.get_random_filename()
    async with aiohttp.ClientSession() as session:
        async with session.get(attachment.url) as response:
            if response.status == 200:
                async with aiofiles.open(file_path, "wb") as f:
                    await f.write(await response.read())
    return file_path


async def save_str_to_file(content: str) -> str:
    file_path = common.get_random_filename()
    async with aiofiles.open(file_path, "w") as f:
        await f.write(content)
    return file_path


@logger.async_try_log()
async def handle_order_item_reaction_question_mark(bot: commands.Bot, user: dc.User):
    help_message = (
        "❔ **Pomoc dotycząca realizacji zamówienia** ❔\n"
        "Kliknij odpowiednią reakcję, aby zmienić status zamówienia:\n"
    )
    for status in mrowka_data.MrowkaOrderItemStatus:
        help_message += f"{status.help_text()}\n"
    help_message += (
        "Aby edytować zamówienie, pobierz załączony plik CSV, wprowadź zmiany i prześlij go z powrotem.\n"
        "Aktualizując zamówienie nie można dodawać nowych butów!\n"
    )
    help_message += f"{common.HELP_PODMIEN_MAIL}\n"
    await user.send(bot, help_message)


@logger.async_try_log()
async def mails_update_csv(bot: commands.Bot, dc_message: discord.Message):
    for attachment in dc_message.attachments:
        if not attachment.filename.lower().startswith("mails") or not attachment.filename.lower().endswith(".csv"):
            continue
        unused_mails_before = await mails.get_number_of_unused_mails()

        file_path = await save_attachment_to_file(attachment)
        try:
            new_mails = await mails.mails_csv_to_mails(file_path)
            await mails.add_new_mails(new_mails)
        except Exception as e:
            import traceback
            error_lines = [line for line in traceback.format_exc().split('\n') if line.strip()]
            short_err = error_lines[-1] if error_lines else str(e)
            message = dc.message_from_dc_message(dc_message)
            await message.author.send(bot, content=f"⚠️ Błąd odczytu pliku upewnij się czy format jest ok CSV **{attachment.filename}**.\nZgłoszony problem: `{short_err}`")
            continue

        unused_mails_after = await mails.get_number_of_unused_mails()
        content = (
            f"✉️ **Dodano {unused_mails_after - unused_mails_before} nowych maili** ✉️\n"
            f"Łącznie dostępnych maili: {unused_mails_after}"
        )
        message = dc.message_from_dc_message(dc_message)
        await message.author.send(bot, content=content)


@logger.async_try_log()
async def tracking_update_csv(bot: commands.Bot, dc_message: discord.Message):
    message = dc.message_from_dc_message(dc_message)
    for attachment in dc_message.attachments:
        if not attachment.filename.lower().startswith("trackings") or not attachment.filename.lower().endswith(".csv"):
            continue

        file_path = await save_attachment_to_file(attachment)
        try:
            async with mrowka_data.PisarzMrowka.lock:
                data = await mrowka_data.PisarzMrowka.read(safe=False)
                new_trackings, fails = await data.update_trackings_from_csv(
                    bot,
                    file_path,
                    message.author,
                )
                await mrowka_data.PisarzMrowka.write(data, safe=False)
        except Exception as e:
            import traceback
            error_lines = [line for line in traceback.format_exc().split('\n') if line.strip()]
            short_err = error_lines[-1] if error_lines else str(e)
            await message.author.send(bot, content=f"⚠️ Błąd odczytu pliku upewnij się czy format jest ok CSV **{attachment.filename}**.\nZgłoszony problem: `{short_err}`")
            continue

        content = f"📦 **Dodano {new_trackings} nowych trackingów** 📦"
        if len(fails) > 0:
            content += "\n❗ **Nie udało się dodać trackingów dla następujących zamówień:** ❗\n"
            for fail in fails:
                content += f"{fail}\n"

        await message.author.send(bot, content=content)


@logger.async_try_log()
async def ticket_update_csv(bot: commands.Bot, message: discord.Message):
    if len(message.attachments) == 0:
        return
    if not hasattr(message.channel, "name"):
        return
    if not isinstance(message.channel.name, str):  # type: ignore
        return

    user = dc.User(
        id=message.author.id,
        name=message.author.display_name,
    )

    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        ticket_name: str = message.channel.name  # type: ignore
        if ticket_name not in data.tickets:
            return
        ticket = data.tickets[ticket_name]

        for attachment in message.attachments:
            if not attachment.filename.lower().endswith(".csv"):
                continue
            if attachment.filename.lower().startswith("mails") or attachment.filename.lower().startswith("trackings"):
                continue

            file_path = await save_attachment_to_file(attachment)
            try:
                shoe_collection = await mrowka_data.MrowkaShoeCollection.from_csv(file_path)
            except Exception as e:
                import traceback
                error_lines = [line for line in traceback.format_exc().split('\n') if line.strip()]
                short_err = error_lines[-1] if error_lines else str(e)
                try:
                    ticket_channel = await dc.channel_from_name(bot, ticket_name)
                    await ticket_channel.send(bot, f"⚠️ Błąd odczytu pliku upewnij się czy format jest ok CSV **{attachment.filename}**.\nZgłoszony problem: `{short_err}`")
                except: pass
                continue

            # FAZA 1: Brakujące ceny (Bloker podziału)
            needs_price_scan = any(s.price == 0 for s in shoe_collection.shoes.values())

            async def run_scanner():
                """Skanuje linki Zalando, aktualizuje ceny i EANy w shoe_collection.
                Usuwa rozmiary bez EAN i notyfikuje użytkownika o usuniętych rozmiarach.
                Po skanie shoe_collection jest gotowy do podziału zamówień."""
                n_links = len(shoe_collection.shoes)
                ticket_channel = await dc.channel_from_name(bot, ticket_name)
                await ticket_channel.send(
                    bot,
                    f"⏳ Skanuję {n_links} produkt(ów) na Zalando — proszę poczekać...\n"
                    f"(ok. {n_links * 20}–{n_links * 40} sek)"
                )
                import zalando_scanner
                import ean_db
                urls = list(shoe_collection.shoes.keys())
                scan_results = await zalando_scanner.scan_urls(urls)

                scan_summary_lines = []
                price_warnings = []
                removed_sizes_info = []  # [(product_name, [removed_sizes])]

                for result in scan_results:
                    s = shoe_collection.shoes.get(result.url)
                    if result.ok:
                        if s:
                            zalando_price = result.current_price
                            csv_price = s.og_price

                            # Walidacja ceny: CSV vs Zalando (ostrzeżenie gdy różnica >5%)
                            if csv_price and zalando_price and abs(csv_price - zalando_price) / zalando_price > 0.05:
                                price_warnings.append(
                                    f"⚠️ **Cena niezgodna** dla **{result.name or result.url[:50]}**\n"
                                    f"   CSV: `{csv_price:.2f} PLN` → Zalando: `{zalando_price:.2f} PLN`\n"
                                    f"   _Używam ceny ze strony Zalando._"
                                )
                            # Zawsze używamy danych z Zalando (w tym SKU)
                            s.og_price = zalando_price
                            s.price = round(zalando_price * 0.8, 2)
                            if result.sku: s.sku = result.sku
                            if result.name: s.name = result.name
                            if result.brand: s.brand = result.brand

                            # Sprawdź które zamówione rozmiary nie mają EAN — ostrzeżenie (nie usuwaj)
                            missing_ean_sizes = [
                                sz for sz in s.size_to_quantity
                                if sz not in result.size_to_ean or not result.size_to_ean.get(sz)
                            ]
                            if missing_ean_sizes:
                                prod_name = result.name or result.url[:50]
                                removed_sizes_info.append((prod_name, missing_ean_sizes))

                        await ean_db.save_scan_results(
                            link=result.url,
                            name=result.name,
                            brand=result.brand,
                            sku=result.sku,
                            price=result.current_price,
                            size_to_ean=result.size_to_ean,
                        )
                        scan_summary_lines.append(
                            f"✅ [{result.brand or '?'}] **{result.name or '?'}** "
                            f"— {result.current_price:.2f} PLN | {len(result.size_to_ean)} EAN"
                        )
                    else:
                        if getattr(result, 'sold_out', False):
                            scan_summary_lines.append(
                                f"🚫 [{result.brand or '?'}] **{result.name or result.url[:50]}** "
                                f"— ARTYKUŁ WYPRZEDANY — pomijam"
                            )
                        else:
                            scan_summary_lines.append(f"❌ {result.url[:60]}... — BŁĄD: {result.error}")

                await ticket_channel.send(bot, f"📊 **Wyniki skanowania:**\n" + "\n".join(scan_summary_lines))

                if price_warnings:
                    await ticket_channel.send(bot, "\n\n".join(price_warnings))

                if removed_sizes_info:
                    lines = [f"{user.mention()} ⚠️ **Rozmiary bez EAN — PZ zostanie utworzone, ale bez kodu kreskowego:**"]
                    for prod_name, missing in removed_sizes_info:
                        lines.append(f"  ⚠️ **{prod_name}**: rozmiary `{', '.join(missing)}` — brak EAN na Zalando")
                    await ticket_channel.send(bot, "\n".join(lines))

                return scan_results  # przekazujemy do zewnętrznego scope
            # Zawsze skanuj Zalando przy nowym zamówieniu — świeże EAN, ceny i SKU
            last_scan_results = await run_scanner() or []

            # === USUŃ WYPRZEDANE PRODUKTY z shoe_collection przed podziałem zamówień ===
            # Produkty bez żadnego accepted_size (OutOfStock) nie powinny trafić ani do
            # zamówień Discord ani do PZ — usuwamy je tutaj raz na zawsze.
            sold_out_urls = {
                r.url for r in last_scan_results
                if getattr(r, 'sold_out', False)
            }
            if sold_out_urls:
                removed_names = []
                for url in sold_out_urls:
                    shoe = shoe_collection.shoes.pop(url, None)
                    if shoe:
                        # MrowkaShoeInfo nie ma .name — użyj nazwy ze skanera lub skróconego linku
                        scan_name = next((r.name for r in last_scan_results if r.url == url and r.name), None)
                        removed_names.append(scan_name or url[:80])
                ticket_channel = await dc.channel_from_name(bot, ticket_name)
                lines = [f"{user.mention()} 🚫 **Usunięto WYPRZEDANE produkty z zamówienia:**"]
                for n in removed_names:
                    lines.append(f"  — ~~{n}~~")
                lines.append("_Produkty te nie trafią do PZ ani do zamówień na Discordzie._")
                await ticket_channel.send(bot, "\n".join(lines))


            # ZAPISZ ZAMÓWIENIA NA DISCORDZIE (po skanie — ceny i rozmiary już poprawne)
            orders_before = set(ticket.divided_orders.keys())
            await ticket.update_divided_orders(bot, shoe_collection, user, data)
            new_order_names = set(ticket.divided_orders.keys()) - orders_before

            # === SKANOWANIE: Upewnij się że produkty istnieją w Subiekcie ===
            # PZ będzie stworzone dopiero po potwierdzeniu zamówienia (ZAMOWIENIE_POTWIERDZONE)
            ticket_channel = await dc.channel_from_name(bot, ticket_name)
            for order_name in new_order_names:
                order_item = ticket.divided_orders[order_name]
                await _ensure_products_for_order_item(bot, ticket_channel, order_item)

            await ticket.send_ticket_csv_message(bot)

        await mrowka_data.PisarzMrowka.write(data, safe=False)




async def _ensure_products_for_order_item(
    bot: commands.Bot,
    ticket_channel,
    order_item: "mrowka_data.MrowkaOrderItem",
) -> None:
    """
    Przy wyśle CSV: skanuje EANy i upewnia się że produkty istnieją w Subiekcie.
    Nie tworzy PZ — PZ jest tworzone przy ZAMOWIENIE_POTWIERDZONE.
    """
    try:
        links = [shoe.link for shoe in order_item.shoes]
        ean_data = await ean_db.get_eans_for_links(links)

        sfera_entries = []
        for shoe_info in order_item.shoes:
            entries_for_shoe = ean_data.get(shoe_info.link, [])
            for size, qty in shoe_info.size_to_quantity.items():
                if qty <= 0:
                    continue
                entry = next((e for e in entries_for_shoe if e.size == size), None)
                if entry and entry.ean:
                    sfera_entries.append({
                        "EAN": entry.ean,
                        "SKU": entry.sku,
                        "Brand": entry.brand,
                        "ModelName": f"{entry.brand} {entry.sku} \u00b7 {entry.size}" if (entry.brand and entry.sku) else (entry.name or f"{entry.brand} \u00b7 {entry.size}"),
                        "Size": entry.size,
                    })

        if sfera_entries:
            ensure_res = await _sfera_cli("EnsureProducts", {"EnsureProducts": sfera_entries}, timeout=180)
            if not ensure_res.get("Success"):
                err_msg = ensure_res.get("Message", "?")
                logger.logger.warning("EnsureProducts (CSV) failed dla %s: %s", order_item.name, err_msg)
                await ticket_channel.send(
                    bot,
                    f"\u26a0\ufe0f Ostrzeżenie: nie udało się zarejestrować produktów w Subiekcie dla **{order_item.name}**.\n"
                    f"Błąd: `{err_msg[:200]}`\n_PZ zostanie spróbowane przy potwierdzeniu zamówienia._",
                )
            else:
                logger.logger.info("EnsureProducts (CSV) OK dla %s: %d prod.", order_item.name, len(sfera_entries))
    except Exception as e:
        logger.logger.exception("_ensure_products_for_order_item: %s", e)


async def _create_pz_for_order_item(
    bot: commands.Bot,
    ticket_channel,
    order_item: "mrowka_data.MrowkaOrderItem",
) -> None:
    """
    Tworzy PZ (Przyjęcie Zewnętrzne) w Subiekcie dla jednego order item.
    Każdy order_item to osobna paczka na Zalando (limit kwotowy).
    Pobiera EANy z ean_db dla linków/rozmiarów przypisanych do TEGO order_item.

    order_item.shoes to lista MrowkaShoeInfo już podzielona przez take_one_order_item.
    Np. duży ticket 20k PLN → 5 order_items → 5 osobnych PZ.
    """
    try:
        # Zbierz linki tylko z butów należących do TEGO order itemu
        links = [shoe.link for shoe in order_item.shoes]
        ean_data = await ean_db.get_eans_for_links(links)

        pz_items = []
        for shoe_info in order_item.shoes:
            for size, qty in shoe_info.size_to_quantity.items():
                if qty <= 0:
                    continue
                entries = ean_data.get(shoe_info.link, [])
                entry = next((e for e in entries if e.size == size), None)
                ean = entry.ean if entry else None
                if not ean:
                    logger.logger.warning(
                        "Brak EAN dla %s r.%s w order %s — pomijam",
                        shoe_info.link[:50], size, order_item.name
                    )
                    continue
                pz_items.append({
                    "ean": ean,
                    "symbol": entry.sku if entry else None,
                    "quantity": qty,
                    "cenaZakupu": round(shoe_info.og_price * 0.8, 2),  # cena zakupu = Zalando × 0.80
                })

        if not pz_items:
            await ticket_channel.send(
                bot,
                f"⚠️ Brak EANów dla **{order_item.name}** — PZ pominięte.\n"
                f"_Produkt może być niedostępny na Zalando lub skaner nie wykrył EANów._\n"
                f"Sprawdź dostępność na stronie i wyślij CSV ponownie jeśli towar wróci.",
            )
            return

        # === ENSURE: Upewnij się że produkty istnieją w Subiekcie przed PZ ===
        sfera_entries = []
        for shoe_info in order_item.shoes:
            entries_for_shoe = ean_data.get(shoe_info.link, [])
            for size, qty in shoe_info.size_to_quantity.items():
                if qty <= 0:
                    continue
                entry = next((e for e in entries_for_shoe if e.size == size), None)
                if entry and entry.ean:
                    sfera_entries.append({
                        "EAN": entry.ean,
                        "SKU": entry.sku,
                        "Brand": entry.brand,
                        # Format nazwy w Subiekcie: "Nike HV4517-100 · 43" (prawdziwe SKU, nie nazwa z Zalando)
                        "ModelName": f"{entry.brand} {entry.sku} \u00b7 {entry.size}" if (entry.brand and entry.sku) else (entry.name or f"{entry.brand} \u00b7 {entry.size}"),
                        "Size": entry.size,
                    })
        if sfera_entries:
            ensure_res = await _sfera_cli("EnsureProducts", {"EnsureProducts": sfera_entries}, timeout=180)
            if not ensure_res.get("Success"):
                err_msg = ensure_res.get("Message", "?")
                logger.logger.warning("EnsureProducts failed dla %s: %s", order_item.name, err_msg)
                await ticket_channel.send(
                    bot,
                    f"⚠️ Nie można stworzyć produktów w Subiekcie dla **{order_item.name}** — PZ pominięte.\n"
                    f"Błąd: `{err_msg[:200]}`",
                )
                return

        mail_email = order_item.mail.mail if order_item.mail else None
        zalando_pass = order_item.mail.zalando_pass if order_item.mail else None

        payload = {

            "dostawcaEmail": None,
            "dostawcaNip": None,
            "dataOperacji": None,
            "uwagi": (
                f"Zamówienie: {order_item.name} | "
                f"Konto: {mail_email or 'nieprzypisane'} | "
                f"Hasło Zalando: {zalando_pass or '—'}"
            ),
            "items": pz_items,
        }

        result = await _sfera_cli("CreatePZ", {"PzData": payload}, timeout=120)
        if result.get("Success") and result.get("DocumentNumber"):
            doc_num = result["DocumentNumber"]
            order_item.pz_sygnatura = doc_num  # zapamiętaj do późniejszej aktualizacji
            await ticket_channel.send(
                bot,
                f"📦 **PZ `{doc_num}`** → zamówienie **{order_item.name}**\n"
                f"Konto: `{mail_email or '?'}` | Pozycji: {len(pz_items)} szt.",
            )
            logger.logger.info("PZ %s created for order %s (%d items)", doc_num, order_item.name, len(pz_items))
            # Odśwież wiadomości Discorda — żeby PZ nr pojawił się w tickecie i magazynie
            try:
                async with mrowka_data.PisarzMrowka.lock:
                    _data = await mrowka_data.PisarzMrowka.read(safe=False)
                    _ticket = _data.tickets.get(order_item.ticket_name)
                    if _ticket:
                        _oi = _ticket.divided_orders.get(order_item.name)
                        if _oi:
                            _oi.pz_sygnatura = doc_num
                            await _oi.discord_update(bot, _data)
                    await mrowka_data.PisarzMrowka.write(_data, safe=False)
            except Exception as _e:
                logger.logger.warning("_create_pz_for_order_item: discord_update po PZ fail: %s", _e)
        else:
            await ticket_channel.send(
                bot,
                f"⚠️ Sfera CLI błąd — PZ dla **{order_item.name}** pominięte.\n"
                f"Szczegóły: `{result.get('Message', '?')}`",
            )

    except Exception as e:
        logger.logger.exception("_create_pz_for_order_item: %s", e)

# ─────────────────────────────────────────────────────────────────────────────

async def remove_cancelled_from_pz(order_item: "mrowka_data.MrowkaOrderItem") -> None:
    """
    Po potwierdzeniu zamówienia odejmuje anulowane ilości z PZ.
    - cancelled_shoes zawiera RÓŻNICĘ (ile zostało odjęte przez magazyniera)
    - Jeśli pozycja w PZ ma np. 28 szt. a anulowano 10 → ustawiamy 18 (nie usuwamy!)
    - Usuwamy pozycję tylko jeśli ilość spada do 0
    """
    if not order_item.cancelled_shoes:
        return  # nic nie anulowano
    if not order_item.pz_sygnatura:
        logger.logger.warning("remove_cancelled_from_pz: brak pz_sygnatura dla %s", order_item.name)
        return

    # Zbierz EAN-y anulowanych rozmiarów z ich ilościami (delta)
    all_links = list({shoe.link for shoe in order_item.cancelled_shoes})
    link_to_entries: dict[str, list] = {}
    for link in all_links:
        link_to_entries[link] = await ean_db.get_eans_for_link(link)

    # {ean: qty_to_subtract}
    ean_delta: dict[str, int] = {}
    for shoe in order_item.cancelled_shoes:
        entries = link_to_entries.get(shoe.link, [])
        for size, qty in shoe.size_to_quantity.items():
            if qty <= 0:
                continue
            entry = next((e for e in entries if e.size == size), None)
            if entry is None or not entry.ean:
                continue
            ean_delta[entry.ean] = ean_delta.get(entry.ean, 0) + qty

    if not ean_delta:
        logger.logger.info("remove_cancelled_from_pz: brak EAN-ów do aktualizacji dla %s", order_item.name)
        return

    items = [{"EAN": ean, "QtyToSubtract": qty} for ean, qty in ean_delta.items()]
    payload = {"PzSygnatura": order_item.pz_sygnatura, "Items": items}

    result = await _sfera_cli("UpdatePzPositions", {"UpdatePzData": payload}, timeout=120)
    if result.get("Success"):
        logger.logger.info(
            "UpdatePzPositions OK dla %s (%s): %s",
            order_item.name, order_item.pz_sygnatura, result.get("Message", "")
        )
    else:
        logger.logger.warning(
            "UpdatePzPositions FAILED dla %s: %s",
            order_item.name, result.get("Message", "?")
        )


# ─────────────────────────────────────────────────────────────────────────────

# Godzinny check IMAP — trigger faktury dostępnej do pobrania
# ─────────────────────────────────────────────────────────────────────────────

MAGAZYNIERZY_CHANNEL = "magazynierzy"
ZADANIA_CATEGORY = "Zadania"
INVOICE_TRIGGER_SUBJECTS = ["faktura", "invoice", "rachunek", "your invoice", "twoja faktura"]


async def send_zadanie_potwierdzenie(
    bot: commands.Bot,
    ticket: "mrowka_data.MrowkaTicket",
    zk_num: Optional[str],
    zk_error: Optional[str] = None,
) -> None:
    """
    Wysyła zadanie do kanału 'zadania-{owner}' w kategorii 'Zadania'.
    Treść: ile par potwierdzono/anulowano, wartość, numer ZK.
    React ✅ usuwa wiadomość (dismiss).
    """
    import re as _re
    import discord as _discord

    owner_name = ticket.owner.name if ticket.owner else "nieznany"
    clean_name = _re.sub(r"[^a-z0-9]", "", owner_name.lower()) or "bot"
    channel_name = f"zadania-{clean_name}"

    # --- Policz pary i wartość ---
    confirmed_pairs = 0
    cancelled_pairs = 0
    confirmed_value = 0.0

    for oi in ticket.divided_orders.values():
        status = oi.history.get_status().status.get_status()
        if status == mrowka_data.MrowkaOrderItemStatusStatus.ANULOWANE:
            # Całe zamówienie anulowane — wszystkie pary do anulowanych
            cancelled_pairs += sum(sum(shoe.size_to_quantity.values()) for shoe in oi.shoes)
        else:
            # Potwierdzone pary
            confirmed_pairs += sum(sum(shoe.size_to_quantity.values()) for shoe in oi.shoes)
            confirmed_value += oi.price_total()
            # Częściowe anulowania wewnątrz potwierdzonego zamówienia
            cancelled_pairs += sum(sum(shoe.size_to_quantity.values()) for shoe in oi.cancelled_shoes)

    # --- Zbuduj wiadomość zadania ---
    zk_line = f"📄 ZK: **{zk_num}**" if zk_num else f"⚠️ ZK: błąd — `{zk_error}`"
    content = (
        f"📋 **Zadanie PVK — {ticket.name}**\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"✅ Potwierdzone: **{confirmed_pairs} par** — {confirmed_value:.2f} PLN\n"
        f"❌ Anulowane: **{cancelled_pairs} par**\n"
        f"{zk_line}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"*Zreaguj ✅ żeby odrzucić*"
    )

    # --- Znajdź/utwórz kanał zadania ---
    zadania_channel = None
    for guild in bot.guilds:
        # Szukaj istniejącego kanału
        zadania_channel = _discord.utils.get(guild.text_channels, name=channel_name)
        if zadania_channel:
            break

        # Stwórz kategorię Zadania jeśli nie ma
        category = _discord.utils.get(guild.categories, name=ZADANIA_CATEGORY)
        if category is None:
            category = await guild.create_category(ZADANIA_CATEGORY, position=0)

        # Stwórz kanał
        zadania_channel = await guild.create_text_channel(channel_name, category=category)
        break

    if zadania_channel is None:
        logger.logger.warning("send_zadanie_potwierdzenie: brak kanału '%s'", channel_name)
        return

    sent = await zadania_channel.send(content)
    await sent.add_reaction("✅")

    # Zarejestruj w data (do dismiss-a przez reakcję)
    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        data.message_id_to_message_type[sent.id] = mrowka_data.MrowkaMessageType.ZADANIE
        await mrowka_data.PisarzMrowka.write(data, safe=False)


@logger.async_try_log()
async def check_imap_invoice(bot: commands.Bot) -> None:
    """
    Sprawdza Gmail IMAP wszystkich kont w poszukiwaniu maili o fakturze.
    Trigger: mail od Zalando z "faktura" / "invoice" w temacie.
    Po znalezieniu → powiadamia magazynierów na Discordzie.
    """
    try:
        import imaplib
        import email as email_mod
        from email.header import decode_header as decode_hdr

        import json, pathlib
        config_path = pathlib.Path(__file__).parent / "config.json"
        config = json.loads(config_path.read_text(encoding="utf-8"))
        accounts = config.get("gmail_accounts", [])

        # Odczytaj aktualny stan ticketów żeby zmapować mail → order_item
        data = await mrowka_data.PisarzMrowka.read()
        mail_to_order: dict[str, "mrowka_data.MrowkaOrderItem"] = {}
        for ticket in data.tickets.values():
            for order_item in ticket.divided_orders.values():
                if order_item.mail:
                    mail_to_order[order_item.mail.mail.lower()] = order_item

        for account in accounts:
            email_addr = account["email"]
            app_password = account["app_password"]
            try:
                mail_conn = imaplib.IMAP4_SSL("imap.gmail.com", 993)
                mail_conn.login(email_addr, app_password)
                mail_conn.select("INBOX", readonly=True)

                _, data_ids = mail_conn.search(None, '(UNSEEN FROM "zalando")')
                ids = data_ids[0].split()

                for num in ids:
                    _, msg_data = mail_conn.fetch(num, "(RFC822.HEADER)")
                    msg = email_mod.message_from_bytes(msg_data[0][1])

                    # Dekoduj temat
                    raw_subj = msg.get("Subject", "")
                    parts = decode_hdr(raw_subj)
                    subject = ""
                    for part, enc in parts:
                        if isinstance(part, bytes):
                            subject += part.decode(enc or "utf-8", errors="ignore")
                        else:
                            subject += str(part)
                    subject_lower = subject.lower()

                    # Sprawdź trigger faktury
                    if not any(kw in subject_lower for kw in INVOICE_TRIGGER_SUBJECTS):
                        continue

                    to_header = msg.get("To", "")
                    # Znajdź adres email z To:
                    import re as _re
                    to_match = _re.search(r"[\w.+%\-]+@[\w.\-]+", to_header)
                    to_email = to_match.group(0).lower() if to_match else email_addr.lower()

                    # Dopasuj do order_item
                    order_item = mail_to_order.get(to_email)

                    logger.logger.info("📧 Trigger faktury: %s → %s", subject[:60], to_email)
                    await _notify_magazynierzy_invoice(bot, order_item, to_email, subject)

                mail_conn.logout()

            except Exception as e:
                logger.logger.error("check_imap_invoice konto %s: %s", email_addr, e)

    except Exception as e:
        logger.logger.exception("check_imap_invoice: %s", e)


@logger.async_try_log()
async def check_gmail_delivery(bot: commands.Bot) -> None:
    """
    Sprawdza Gmail IMAP wszystkich kont w poszukiwaniu maili dostawczych
    ('Twoja przesyłka zostanie dostarczona') i wyciąga tracking + kwotę.

    Logika podwójnej paczki:
      - Przy pierwszym mailu: jeśli |kwota_paczki - pz_total| <= 10 PLN
        → jedna paczka → od razu ZAMOWIENIE_WYSLANE
      - Jeśli różnica > 10 PLN → split shipment (max 2 paczki)
        → czekaj na drugą paczkę, WYSLANE dopiero po niej
      - Po każdej paczce: aktualizuj uwagi PZ o tracking i kwotę

    Tylko UNSEEN maile — po przetworzeniu oznacza jako przeczytane.
    """
    SPLIT_TOLERANCE_PLN = 10.0  # max roznica miedzy kwota paczki a total PZ

    try:
        deliveries = await asyncio.to_thread(gmail_imap.get_all_new_delivery_emails)
        if not deliveries:
            logger.logger.info("check_gmail_delivery: brak nowych maili dostawczych")
            return

        async with mrowka_data.PisarzMrowka.lock:
            data = await mrowka_data.PisarzMrowka.read(safe=False)
            mail_to_order: dict[str, "mrowka_data.MrowkaOrderItem"] = {}
            for ticket in data.tickets.values():
                for oi in ticket.divided_orders.values():
                    if oi.mail:
                        mail_to_order[oi.mail.mail.lower()] = oi

            changed = False
            for info in deliveries:
                if not info.tracking:
                    logger.logger.info(
                        "check_gmail_delivery: brak trackingu w mailu dla %s",
                        info.zalando_account
                    )
                    continue

                logger.logger.info(
                    "check_gmail_delivery: tracking=%s, konto=%s, kwota=%s PLN",
                    info.tracking, info.zalando_account, info.shipping_amount
                )

                oi = mail_to_order.get(info.zalando_account.lower())

                # 1. Update PZ tracking w Subiekcie
                if oi:
                    order_name = oi.name
                    result = await _subiekt_post("/api/pz/update-tracking", {
                        "orderName": order_name,
                        "tracking": info.tracking,
                    })
                    if result and result.get("Success"):
                        logger.logger.info("update-tracking OK: %s -> %s", order_name, info.tracking)
                    else:
                        logger.logger.warning("update-tracking failed: orderName=%s", order_name)
                else:
                    logger.logger.info(
                        "check_gmail_delivery: konto %s nie pasuje do zadnego aktywnego zamowienia",
                        info.zalando_account
                    )

                if not oi:
                    continue

                # 2. Zapisz tracking i date dostawy w order_item
                oi.tracking = info.tracking
                if info.delivery_date:
                    oi.delivery_date = info.delivery_date
                if info.order_number:
                    oi.order_number = info.order_number

                # 3. Dodaj ShipmentInfo TYLKO z maila Zalando (ma order_number).
                # Mail InPost dotyczy tej samej paczki co mail Zalando — gdybysmy
                # dodali z obu, liczylibysmy jedna paczke dwa razy (np. 3100 + 3100).
                # Wyroznik: Zalando mail ma order_number != None, InPost ma order_number = None.
                is_zalando_mail = info.order_number is not None

                if is_zalando_mail:
                    # Deduplikacja po kwocie: jesli mamy juz ShipmentInfo z ta sama kwota
                    # (tolerancja 1 PLN), nie dodajemy duplikatu
                    existing_amounts = [s.amount_pln for s in oi.shipments if s.amount_pln is not None]
                    amount_already_counted = any(
                        abs((info.shipping_amount or 0) - a) <= 1.0
                        for a in existing_amounts
                    ) if info.shipping_amount and existing_amounts else False

                    existing_trackings = {s.tracking for s in oi.shipments}

                    if info.tracking not in existing_trackings and not amount_already_counted:
                        import datetime as _dt
                        shipment = mrowka_data.ShipmentInfo(
                            tracking=info.tracking,
                            amount_pln=info.shipping_amount,
                            date_sent=_dt.datetime.now().strftime("%Y-%m-%d"),
                            source_mail=info.gmail_base,
                        )
                        oi.shipments.append(shipment)
                        logger.logger.info(
                            "check_gmail_delivery: [Zalando] dodano ShipmentInfo dla %s "
                            "(kwota=%.2f PLN, paczka #%d)",
                            oi.name,
                            info.shipping_amount or 0.0,
                            len(oi.shipments),
                        )
                    elif amount_already_counted:
                        logger.logger.info(
                            "check_gmail_delivery: [Zalando] pominiety duplikat ShipmentInfo dla %s "
                            "(kwota=%.2f juz jest w liscie)",
                            oi.name, info.shipping_amount or 0.0
                        )
                else:
                    logger.logger.info(
                        "check_gmail_delivery: [InPost] mail dla %s — tracking=%s "
                        "(nie dodaje ShipmentInfo, ta paczka juz liczona z maila Zalando)",
                        oi.name, info.tracking
                    )

                # 4. Aktualizuj uwagi PZ o przesylkach (przez sygnature PZ, jesli znana)
                uwagi_shipments = oi.shipments_uwagi()
                if uwagi_shipments and oi.pz_sygnatura:
                    upd = await _subiekt_post("/api/pz/update-uwagi", {
                        "PzSygnatura": oi.pz_sygnatura,
                        "Uwagi": uwagi_shipments,
                    })
                    if upd and upd.get("Success"):
                        logger.logger.info(
                            "check_gmail_delivery: uwagi PZ zaktualizowane dla %s (%s): %s",
                            oi.name, oi.pz_sygnatura, uwagi_shipments
                        )
                    else:
                        logger.logger.warning(
                            "check_gmail_delivery: blad zapisu uwagi PZ %s: %s",
                            oi.pz_sygnatura, upd
                        )
                elif uwagi_shipments:
                    logger.logger.info(
                        "check_gmail_delivery: brak pz_sygnatura dla %s — pomijam zapis uwagi shipments",
                        oi.name
                    )

                # 5. Sprawdz czy oznaczyc jako WYSLANE
                cur = oi.history.get_status()
                wyslane = mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_WYSLANE
                zrealizowane = mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_ZOSTALO_ZREALIZOWANE

                if cur.status in (wyslane, zrealizowane):
                    continue  # juz wyslane, nie zmieniamy

                pz_total = oi.price_total()
                n_shipments = len(oi.shipments)

                # Sprawdz czy to jedna czy dwie paczki na podstawie PIERWSZEJ paczki
                first_amount = oi.shipments[0].amount_pln if oi.shipments else None

                if first_amount is not None and pz_total > 0:
                    diff = abs(first_amount - pz_total)
                    is_single = diff <= SPLIT_TOLERANCE_PLN
                else:
                    # Brak kwoty — traktuj jako jedna paczka (bezpieczny default)
                    is_single = True

                should_mark_wyslane = False
                if is_single:
                    # Jedna paczka — od razu wyslane
                    should_mark_wyslane = True
                    logger.logger.info(
                        "check_gmail_delivery: %s — JEDNA paczka (kwota=%.2f, total=%.2f, diff=%.2f) -> WYSLANE",
                        oi.name, first_amount or 0, pz_total,
                        abs((first_amount or 0) - pz_total)
                    )
                else:
                    # Split shipment — czekaj az beda 2 paczki
                    if n_shipments >= 2:
                        should_mark_wyslane = True
                        logger.logger.info(
                            "check_gmail_delivery: %s — DWIE paczki odebrane -> WYSLANE "
                            "(kwota1=%.2f, kwota2=%.2f, total=%.2f)",
                            oi.name,
                            oi.shipments[0].amount_pln or 0,
                            oi.shipments[1].amount_pln or 0,
                            pz_total,
                        )
                    else:
                        # Pierwsza z dwoch paczek — powiadom Discord
                        logger.logger.info(
                            "check_gmail_delivery: %s — SPLIT SHIPMENT, pierwsza paczka "
                            "(kwota=%.2f, total=%.2f, brakuje=%.2f PLN). Czekam na druga.",
                            oi.name, first_amount or 0, pz_total,
                            pz_total - (first_amount or 0)
                        )
                        # Wyslij powiadomienie na magazynierzy ze to split
                        try:
                            magazyn_channel = None
                            for guild in bot.guilds:
                                for ch in guild.text_channels:
                                    if ch.name == MAGAZYNIERZY_CHANNEL:
                                        magazyn_channel = ch
                                        break
                                if magazyn_channel:
                                    break
                            if magazyn_channel:
                                await magazyn_channel.send(
                                    f"📦 **Split shipment** — zamówienie `{oi.name}`\n"
                                    f"Pierwsza paczka: `{info.tracking}` — {first_amount:.2f} PLN\n"
                                    f"Wartość PZ: **{pz_total:.2f} PLN**\n"
                                    f"Brakuje: **{pz_total - first_amount:.2f} PLN** (druga paczka w drodze)"
                                )
                        except Exception as _e:
                            logger.logger.warning("check_gmail_delivery: blad powiadomienia split: %s", _e)
                        changed = True  # zapisz stan (dodano ShipmentInfo)
                        # Odswierz wiadomosci Discord - pokaz kwote 3100/3900 PLN
                        await oi.discord_update(bot, data)
                        continue

                if should_mark_wyslane:
                    await oi.change_status(bot, wyslane, cur.user, data)
                    changed = True
                    logger.logger.info(
                        "check_gmail_delivery: %s -> ZAMOWIENIE_WYSLANE (tracking=%s)",
                        oi.name, info.tracking
                    )

            if changed:
                await mrowka_data.PisarzMrowka.write(data, safe=False)

    except Exception as e:
        logger.logger.exception("check_gmail_delivery: %s", e)


@logger.async_try_log()
async def check_gmail_delivery_confirmed(bot: commands.Bot) -> None:
    """
    Skanuje IMAP w poszukiwaniu maili 'Cyfrowy dowod dostawy' / 'Przesylka zostala dostarczona'.

    Logika split shipment:
      - expected = max(len(oi.shipments), 1)
      - WYSLANE + faktura request na Discord dopiero gdy delivery_confirmations >= expected
      - ZREALIZOWANE triggerowane osobno w handle_faktura_status_reaction po 'faktura = True'

    Mapowanie: po koncie Zalando (To: header -> oi.mail.mail)
    """
    try:
        confirmed_list = await asyncio.to_thread(gmail_imap.get_all_new_delivery_confirmed_emails)
        if not confirmed_list:
            logger.logger.info("check_gmail_delivery_confirmed: brak nowych maili")
            return

        logger.logger.info("check_gmail_delivery_confirmed: %d maili do przetworzenia", len(confirmed_list))

        async with mrowka_data.PisarzMrowka.lock:
            data = await mrowka_data.PisarzMrowka.read(safe=False)

            mail_to_oi: dict[str, mrowka_data.MrowkaOrderItem] = {}
            for ticket in data.tickets.values():
                for oi in ticket.divided_orders.values():
                    if oi.mail:
                        mail_to_oi[oi.mail.mail.lower()] = oi

            changed = False
            wyslane   = mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_WYSLANE
            zrealizowane = mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_ZOSTALO_ZREALIZOWANE

            for info in confirmed_list:
                account = info.zalando_account.lower()
                oi = mail_to_oi.get(account)
                if oi is None:
                    logger.logger.warning(
                        "check_gmail_delivery_confirmed: nie znaleziono order_item dla konta %s", account
                    )
                    continue

                cur = oi.history.get_status()

                # Juz zrealizowane — pomijamy
                if cur.status == zrealizowane:
                    logger.logger.info(
                        "check_gmail_delivery_confirmed: %s juz ZREALIZOWANE — pomijam", oi.name
                    )
                    continue

                # Dedup — sprawdz czy ten mail juz przetwarzalismy
                # Klucz: order_number (jesli dostepny) lub account (fallback)
                dedup_key = info.order_number or account
                if dedup_key in oi.confirmed_delivery_keys:
                    logger.logger.info(
                        "check_gmail_delivery_confirmed: %s — duplikat maila (key=%s), pomijam",
                        oi.name, dedup_key
                    )
                    continue
                oi.confirmed_delivery_keys.add(dedup_key)

                # Zlicz potwierdzenie dostawy
                oi.delivery_confirmations += 1
                changed = True

                expected = max(len(oi.shipments), 1)

                logger.logger.info(
                    "check_gmail_delivery_confirmed: %s — potwierdzenie %d/%d (order=%s)",
                    oi.name, oi.delivery_confirmations, expected, info.order_number
                )

                # Rejestruj w sesji rush i uruchom rush mode jesli pierwszy mail dnia
                _rush_register_delivery(oi.name)
                _maybe_start_rush_mode(bot)

                if oi.delivery_confirmations < expected:
                    # Split shipment — czekamy na kolejna paczke, tylko update Discord
                    logger.logger.info(
                        "check_gmail_delivery_confirmed: %s — split, czekam %d/%d",
                        oi.name, oi.delivery_confirmations, expected
                    )
                    await oi.discord_update(bot, data)
                    continue

                # Wszystkie paczki dostarczone — zapewnij status WYSLANE
                # (jesli status byl nizszy niz WYSLANE, przeskocz do WYSLANE)
                if wyslane in cur.status.next_statuses() or cur.status == wyslane:
                    if cur.status != wyslane:
                        await oi.change_status(bot, wyslane, cur.user, data)
                        logger.logger.info(
                            "check_gmail_delivery_confirmed: %s → WYSLANE (skip z %s)",
                            oi.name, cur.status.name
                        )
                    else:
                        # Juz WYSLANE — tylko odswiezamy Discord (ponowi faktura request jesli nie bylo)
                        await oi.discord_update(bot, data)
                else:
                    # Status jest juz za WYSLANE albo anomalia — tylko discord_update
                    await oi.discord_update(bot, data)

                # Jezeli faktura juz zatwierdzona (np. przyszla wczesniej) — od razu ZREALIZOWANE
                if oi.faktura is True:
                    await _trigger_zrealizowane(bot, oi, data, info.tracking)

            if changed:
                await mrowka_data.PisarzMrowka.write(data, safe=False)

    except Exception as e:
        logger.logger.exception("check_gmail_delivery_confirmed: %s", e)


async def _trigger_zrealizowane(
    bot: "commands.Bot",
    oi: "mrowka_data.MrowkaOrderItem",
    data: "mrowka_data.MrowkaData",
    tracking: Optional[str] = None,
) -> None:
    """
    Przejdz do ZREALIZOWANE i wywolaj /api/pz/accept.
    Wywolywane gdy: delivery_confirmations >= expected AND faktura == True.
    """
    zrealizowane = mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_ZOSTALO_ZREALIZOWANE
    cur = oi.history.get_status()
    if cur.status == zrealizowane:
        return
    if zrealizowane not in cur.status.next_statuses():
        logger.logger.warning(
            "_trigger_zrealizowane: %s — status %s nie pozwala na ZREALIZOWANE",
            oi.name, cur.status.name
        )
        return
    await oi.change_status(bot, zrealizowane, cur.user, data)
    logger.logger.info("_trigger_zrealizowane: %s → ZREALIZOWANE", oi.name)

    # Przyjecie w Subiekcie — uzywamy trackingu z biezacego maila
    # lub ostatniego trackingu z ListyShipments
    pz_tracking = tracking
    if not pz_tracking and oi.shipments:
        pz_tracking = oi.shipments[-1].tracking
    if pz_tracking:
        pz_result = await _subiekt_post(
            "/api/pz/accept", {"tracking": pz_tracking}, timeout=30
        )
        if pz_result and pz_result.get("sygnatura"):
            logger.logger.info(
                "_trigger_zrealizowane: PZ accept OK, sygnatura=%s", pz_result["sygnatura"]
            )


@logger.async_try_log()
async def check_gmail_delay(bot: commands.Bot) -> None:
    """
    Sprawdza Gmail IMAP wszystkich kont w poszukiwaniu maili o opoznieniu dostawy
    ('Aktualizacja dotyczaca Twojej dostawy' + tresc o opoznieniu).

    Po znalezieniu:
      1. Wysyla powiadomienie na kanal #magazynierzy
      2. Loguje order_item jesli pasuje do konta
    Tylko UNSEEN - po przetworzeniu oznacza jako przeczytane.
    """
    try:
        delays = await asyncio.to_thread(gmail_imap.get_all_new_delay_emails)
        if not delays:
            logger.logger.info("check_gmail_delay: brak nowych maili opoznienia")
            return

        # Znajdz kanal #magazynierzy
        magazyn_channel = None
        for guild in bot.guilds:
            for ch in guild.text_channels:
                if ch.name == MAGAZYNIERZY_CHANNEL:
                    magazyn_channel = ch
                    break
            if magazyn_channel:
                break

        # Mapowanie mail → order_item
        data = await mrowka_data.PisarzMrowka.read()
        mail_to_order: dict[str, "mrowka_data.MrowkaOrderItem"] = {}
        for ticket in data.tickets.values():
            for oi in ticket.divided_orders.values():
                if oi.mail:
                    mail_to_order[oi.mail.mail.lower()] = oi

        for info in delays:
            logger.logger.info("Opoznienie dostawy: konto=%s", info.zalando_account)

            oi = mail_to_order.get(info.zalando_account.lower())
            order_name = oi.name if oi else "nieznane"

            msg = (
                f"\u26a0\ufe0f **Opoznienie dostawy!**\n"
                f"\U0001f4e7 Konto: `{info.zalando_account}`\n"
                f"\U0001f4e6 Zamowienie: `{order_name}`\n"
                f"\U0001f4e9 Zalando: _{info.subject}_\n\n"
                f"_Paczka dotrze z opoznieniem \u2014 brak akcji wymagany, "
                f"bot zaktualizuje status gdy dotrze._"
            )

            if magazyn_channel:
                await magazyn_channel.send(msg)
            else:
                logger.logger.warning("check_gmail_delay: brak kanalu '%s'", MAGAZYNIERZY_CHANNEL)

    except Exception as e:
        logger.logger.exception("check_gmail_delay: %s", e)


async def _notify_magazynierzy_invoice(
    bot: commands.Bot,
    order_item: Optional["mrowka_data.MrowkaOrderItem"],
    email_konta: str,
    mail_subject: str,
) -> None:
    """Wysyła powiadomienie do kanału magazynierzy o dostępnej fakturze."""
    try:
        magazyn_channel = None
        for guild in bot.guilds:
            for ch in guild.text_channels:
                if ch.name == MAGAZYNIERZY_CHANNEL:
                    magazyn_channel = ch
                    break
            if magazyn_channel:
                break

        if magazyn_channel is None:
            logger.logger.warning("Nie znaleziono kanału '%s'", MAGAZYNIERZY_CHANNEL)
            return

        # Pobierz hasło Zalando z banku maili
        zalando_pass = "—"
        if order_item and order_item.mail:
            zalando_pass = order_item.mail.zalando_pass or "—"
        else:
            # Spróbuj znaleźć w banku maili po emailu
            all_mails = await mails.get_all_mails()
            found = next((m for m in all_mails if m.mail.lower() == email_konta.lower()), None)
            if found:
                zalando_pass = found.zalando_pass or "—"

        order_name = order_item.name if order_item else "nieznane"

        content = (
            f"🧾 **Faktura gotowa do pobrania!**\n"
            f"📦 Zamówienie: `{order_name}`\n"
            f"📧 Konto: `{email_konta}`\n"
            f"🔑 Hasło Zalando: `{zalando_pass}`\n\n"
            f"**Kroki:**\n"
            f"1. Zaloguj się na [zalando.pl](https://www.zalando.pl) na konto `{email_konta}`\n"
            f"2. Wejdź w Moje zamówienia → Pobierz fakturę (PDF)\n"
            f"3. **Odeślij plik PDF faktury** tutaj ↩️\n\n"
            f"_(Temat maila: {mail_subject[:80]})_"
        )

        await magazyn_channel.send(content)
        logger.logger.info("Powiadomiono magazynierzy o fakturze dla %s", email_konta)

    except Exception as e:
        logger.logger.exception("_notify_magazynierzy_invoice: %s", e)


@logger.async_try_log()
async def handle_faktura_pdf(bot: commands.Bot, message: discord.Message) -> None:
    """
    Nasłuchuje na PDF faktury Zalando.
    Dwa tryby:
      A) REPLY: odpowiedź na wiadomość order itemu (w dowolnym kanale) — bot
         identyfikuje order_item po Discord message ID referencji.
      B) DIRECT: wrzucenie PDF na #magazynierzy — bot parsuje nr zamówienia z PDFa.
    """
    if not message.attachments:
        return

    # Sprawdź czy jest PDF Zalando (Invoice*.pdf)
    pdf_attachment = None
    for att in message.attachments:
        if att.filename.lower().endswith(".pdf") and att.filename.lower().startswith("invoice"):
            pdf_attachment = att
            break
    if pdf_attachment is None:
        return

    try:
        import invoice_parser
    except ImportError:
        logger.logger.error("invoice_parser nie znaleziony — pip install pdfplumber")
        return

    # Pobierz PDF
    file_path = await save_attachment_to_file(pdf_attachment)
    logger.logger.info("Pobrano fakturę PDF: %s → %s", pdf_attachment.filename, file_path)

    # Parsuj PDF (potrzebujemy nr FV i daty zawsze)
    inv = invoice_parser.parse_invoice_pdf(file_path)

    # ─── TRYB A: Reply na wiadomość order itemu ───────────────────────────────
    order_name = None
    reply_mode = bool(message.reference and message.reference.message_id)

    if reply_mode:
        ref_msg_id = message.reference.message_id
        data = await mrowka_data.PisarzMrowka.read()
        order_name = data.message_id_to_order_item_name.get(ref_msg_id)

        if not order_name:
            await message.reply(
                "⚠️ Nie znalazłem order itemu powiązanego z tą wiadomością.\n"
                "Upewnij się że odpowiadasz na wiadomość bota z danymi zamówienia."
            )
            return

    # ─── TRYB B: Direct na #magazynierzy ──────────────────────────────────────
    else:
        if not hasattr(message.channel, "name") or message.channel.name != MAGAZYNIERZY_CHANNEL:
            return  # nie nasz kanał, ignoruj

        if not inv.ok:
            await message.channel.send(
                f"⚠️ Nie udało się sparsować faktury **{pdf_attachment.filename}**\n"
                f"_(NumerFV={inv.invoice_number}, NrZam={inv.order_number})_\n"
                f"Sprawdź plik ręcznie lub odpowiedz PDF-em na wiadomość zamówienia."
            )
            return

        # Znajdź order_item po numerze zamówienia z PDFa
        data = await mrowka_data.PisarzMrowka.read()
        for ticket in data.tickets.values():
            for oi in ticket.divided_orders.values():
                if oi.order_number and oi.order_number == inv.order_number:
                    order_name = oi.name
                    break
            if order_name:
                break

        if not order_name:
            await message.channel.send(
                f"⚠️ Nie znaleziono zamówienia `{inv.order_number}` z faktury.\n"
                f"Spróbuj odpowiadając PDF-em bezpośrednio na wiadomość zamówienia."
            )
            return

    # ─── Wspólna cześć: update PZ + utwórz FZ ────────────────────────────────
    reply_target = message  # odpowiadamy w tym samym miejscu

    if not inv.ok and reply_mode:
        await reply_target.reply(
            f"⚠️ Nie udało się sparsować faktury **{pdf_attachment.filename}**\n"
            f"_(NumerFV={inv.invoice_number}, Data={inv.invoice_date})_"
        )
        return

    msg_parsowana = await reply_target.reply(
        f"\u2705 Faktura **{pdf_attachment.filename}** sparsowana:\n"
        f"\U0001f4c4 Nr FV: `{inv.invoice_number}`\n"
        f"\U0001f4c5 Data: `{inv.invoice_date}`\n"
        f"\U0001f6cd Order: `{order_name}`\n"
        f"\U0001f504 Aktualizuj\u0119 PZ i tworz\u0119 FZ..."
    )

    # Pobierz order_item po order_name (do trackingu)
    _oi = None
    if order_name:
        _data = data if 'data' in locals() else await mrowka_data.PisarzMrowka.read()
        for _t in _data.tickets.values():
            if order_name in _t.divided_orders:
                _oi = _t.divided_orders[order_name]
                break

    # 4. Update PZ — NumerZewnetrzny + DataFakturyDostawcy + Tracking w Uwagach
    upd = await _subiekt_post("/api/pz/update-invoice", {
        "orderName": order_name,
        "invoiceNumber": inv.invoice_number,
        "invoiceDate": inv.invoice_date,
        "tracking": _oi.tracking if _oi else None,
    })
    if upd:
        logger.logger.info("\u2705 PZ zaktualizowany: NrZew=%s, Data=%s | %s", inv.invoice_number, inv.invoice_date, upd.get("Message", "?"))
    else:
        logger.logger.warning("\u26a0\ufe0f update-invoice nie powiod\u0142o si\u0119 dla %s", order_name)

    # 5. Utwórz FZ na podstawie istniejącego PZ przez PzSygnatura (niezawodne)
    # Używamy CreateFZByPz (alias TestFZ) — szuka PZ po sygnaturze, nie po Uwagach
    pz_sig = _oi.pz_sygnatura if _oi else None
    if not pz_sig:
        logger.logger.warning("handle_faktura_pdf: brak pz_sygnatura dla %s — fallback na orderName", order_name)
    fz_result = await _sfera_cli("CreateFZByPz", {
        "CreateFzByPz": {
            "PzSygnatura": pz_sig or order_name,
            "InvoiceDate": inv.invoice_date,
            "InvoiceNumber": inv.invoice_number,
        }
    }, timeout=90)
    if fz_result and (fz_result.get("Success") or fz_result.get("documentNumber")):
        fz_num = fz_result.get("DocumentNumber") or fz_result.get("documentNumber", "?")
        await reply_target.reply(
            f"\U0001f9fe **FZ `{fz_num}` utworzone!**\n"
            f"NrZewnetrzny: `{inv.invoice_number}` | Order: `{order_name}`"
        )

        # Trigger ZREALIZOWANE jesli wszystkie paczki dostarczone
        async with mrowka_data.PisarzMrowka.lock:
            _d2 = await mrowka_data.PisarzMrowka.read(safe=False)
            _oi2 = None
            for _t2 in _d2.tickets.values():
                if order_name in _t2.divided_orders:
                    _oi2 = _t2.divided_orders[order_name]
                    break
            if _oi2 is not None:
                expected = max(len(_oi2.shipments), 1)
                if _oi2.delivery_confirmations >= expected:
                    await _trigger_zrealizowane(bot, _oi2, _d2)
                    await mrowka_data.PisarzMrowka.write(_d2, safe=False)
                    logger.logger.info(
                        "handle_faktura_pdf: %s → ZREALIZOWANE po FZ", order_name
                    )
    else:
        err = fz_result.get("Message", "?") if fz_result else "brak odpowiedzi"
        await reply_target.reply(
            f"\u26a0\ufe0f FZ nie zosta\u0142o utworzone.\n"
            f"Utw\u00f3rz FZ r\u0119cznie: Nr={inv.invoice_number}, Data={inv.invoice_date}\n"
            f"B\u0142\u0105d: `{err[:200]}`"
        )



@logger.async_try_log()
async def handle_label(bot: commands.Bot, message: discord.Message) -> None:
    """
    Nasłuchuje na kanale #magazynierzy na numer tracking wpisany jako wiadomość
    (np. z czytnika kodów kreskowych lub ręcznie).
    
    Gdy magazynier wpisuje/skanuje tracking → bot:
      1. Rozpoznaje że wiadomość to sam numer tracking (15-30 cyfr)
      2. POST /api/pz/accept → StatusDokumentuId=20 (Przyjęty towar), DataMagazynowa=dziś
      3. Szuka order_item po tracking → zmienia status na ZAMOWIENIE_ZOSTALO_ZREALIZOWANE
      4. Potwierdza na kanale
    """
    if not hasattr(message.channel, "name"):
        return
    if message.channel.name != MAGAZYNIERZY_CHANNEL:
        return

    import re as _re

    # Wykryj czy wiadomość to sam numer tracking (15-30 cyfr, opcjonalnie spacje/myślniki)
    content = message.content.strip().replace(" ", "").replace("-", "")
    tracking_match = _re.fullmatch(r"\d{15,30}", content)
    if not tracking_match:
        return  # Nie tracking — ignoruj

    tracking = content

    # === Krok 1: POST /api/pz/accept → zmiana statusu PZ ===
    accept_result = await _subiekt_post("/api/pz/accept", {"tracking": tracking}, timeout=30)
    if accept_result:
        sygnatura = accept_result.get("sygnatura", "?")
        await message.channel.send(
            f"✅ **{sygnatura}** przyjęte!\n"
            f"Tracking: `{tracking}` | PZ → Przyjęty towar 📦"
        )
    else:
        await message.channel.send(
            f"⚠️ Nie znaleziono PZ z trackingiem `{tracking}`\n"
            f"Sprawdź czy paczka była zamówiona przez Mrówkę."
        )
        return

    # === Krok 2: Zmień status Discord order_item ===
    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        found_oi = None
        for ticket in data.tickets.values():
            for oi in ticket.divided_orders.values():
                if oi.tracking == tracking:
                    found_oi = oi
                    break
            if found_oi:
                break

        if found_oi:
            current_status = found_oi.history.get_status()
            if (
                mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_ZOSTALO_ZREALIZOWANE
                in current_status.status.next_statuses()
            ):
                await found_oi.change_status(
                    bot,
                    mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_ZOSTALO_ZREALIZOWANE,
                    current_status.user,
                    data,
                )
                await mrowka_data.PisarzMrowka.write(data, safe=False)
                logger.logger.info(
                    "handle_label: %s → ZAMOWIENIE_ZOSTALO_ZREALIZOWANE", found_oi.name
                )
            else:
                logger.logger.warning(
                    "handle_label: %s nie może przejść do ZREALIZOWANE (status=%s)",
                    found_oi.name, current_status.status
                )
        else:
            logger.logger.warning(
                "handle_label: nie znaleziono order_item dla tracking=%s", tracking
            )


async def order_item_update_csv(bot: commands.Bot, dc_message: discord.Message):
    if len(dc_message.attachments) == 0:
        return
    message = dc.message_from_dc_message(dc_message)

    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        if not hasattr(dc_message.channel, "name") or dc_message.channel.name in data.tickets:  # type: ignore
            return

        for attachment in dc_message.attachments:
            if (
                not attachment.filename.endswith(".csv")
                or attachment.filename == "mails.csv"
                or attachment.filename == "trackings.csv"
            ):
                continue

            order_item_name = attachment.filename.rsplit(".", 1)[0]
            ticket_name = mrowka_data.order_item_name_to_ticket_name(order_item_name)
            if ticket_name not in data.tickets:
                continue

            ticket = data.tickets[ticket_name]
            if order_item_name not in ticket.divided_orders:
                continue

            file_path = await save_attachment_to_file(attachment)
            shoe_collection = await mrowka_data.MrowkaShoeCollection.from_csv(file_path)
            order_item = ticket.divided_orders[order_item_name]
            error_msg = await order_item.update_shoes(
                bot, shoe_collection, message.author, data
            )
            if error_msg is not None:
                await message.author.send(bot, content=error_msg)

        await mrowka_data.PisarzMrowka.write(data, safe=False)


@logger.async_try_log()
async def handle_order_item_change_state(
    bot: commands.Bot,
    message: discord.Message,
    state: mrowka_data.MrowkaOrderItemStatus,
    user: dc.User,
):
    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        ticket_name = data.message_id_to_ticket_name[message.id]
        order_item_name = data.message_id_to_order_item_name[message.id]
        order_item = data.tickets[ticket_name].divided_orders[order_item_name]
        await order_item.change_status(bot, state, user, data)
        await mrowka_data.PisarzMrowka.write(data, safe=False)


async def handle_order_item_reaction(
    bot: commands.Bot,
    message: discord.Message,
    user: dc.User,
    emoji: str,
):
    if emoji == "❔":
        await handle_order_item_reaction_question_mark(bot, user)
    for status in mrowka_data.MrowkaOrderItemStatus:
        if emoji == status.emoji():
            await handle_order_item_change_state(bot, message, status, user)


@logger.async_try_log()
async def handle_interia_error_reaction(bot, message: dc.Message):
    await message.delete(bot)


@logger.async_try_log()
async def handle_faktura_status_reaction(bot, message: dc.Message, emoji: str):
    """
    Reakcje na wiadomosc faktury - TYLKO wizualne znaczniki.
    NIE triggeruja ZREALIZOWANE. NIE usuwaja wiadomosci.
    Jedynym triggerem ZREALIZOWANE jest odeslanie faktury PDF (handle_faktura_pdf).
    👍🏿 = ok/w trakcie, 👎🏿 = problem
    """
    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        ticket_name = data.message_id_to_ticket_name.get(message.id)
        order_item_name = data.message_id_to_order_item_name.get(message.id)
        if not ticket_name or not order_item_name:
            return
        order_item = data.tickets[ticket_name].divided_orders[order_item_name]

        if emoji == "👍🏿":
            order_item.faktura = True    # status wizualny "ok / w trakcie"
        elif emoji == "👎🏿":
            order_item.faktura = False   # status wizualny "problem"

        # Tylko aktualizacja Discord - bez triggera ZREALIZOWANE
        await order_item.discord_update(bot, data)
        await mrowka_data.PisarzMrowka.write(data, safe=False)


@logger.async_try_log()
async def handle_ean_message_reaction(bot, message: dc.Message):
    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        ticket_name = data.message_id_to_ticket_name[message.id]
        ticket = data.tickets[ticket_name]

    await message.delete(bot)
    await ticket.send_ean_plik_wgrany(bot)


@logger.async_try_log()
async def check_interia(bot: commands.Bot):
    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        ticket_to_order_items = {
            ticket.name: list(ticket.divided_orders.values())
            for ticket in data.tickets.values()
        }

    for ticket_name in ticket_to_order_items.keys():
        for _order_item in ticket_to_order_items[ticket_name]:
            if _order_item is None:
                continue
            if _order_item.mail is None:
                continue
            if (
                _order_item.history.get_status().status
                != mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_POTWIERDZONE
            ):
                if all(
                    [
                        _order_item.tracking is None,
                        _order_item.delivery_date is None,
                        _order_item.order_number is None,
                        _order_item.name_surname is None,
                    ]
                ) or all(
                    [
                        _order_item.tracking is not None,
                        _order_item.delivery_date is not None,
                        _order_item.order_number is not None,
                        _order_item.name_surname is not None,
                    ]
                ):
                    continue

            info = await asyncio.to_thread(
                gmail_imap.get_delivery_info_for_mail,
                _order_item.mail,
            )
            async with mrowka_data.PisarzMrowka.lock:
                data = await mrowka_data.PisarzMrowka.read(safe=False)
                ticket = data.tickets.get(ticket_name, None)
                if ticket is None:
                    continue
                order_item = ticket.divided_orders.get(_order_item.name, None)
                if order_item is None:
                    continue
                if order_item.mail is None:
                    continue
                if info is None:
                    order_item.mail
                    channel_interia = await dc.CHANNEL_INTERIA(bot)
                    message = await channel_interia.send(
                        bot,
                        f"❗ Błąd pobierania informacji z Interii dla konta: {order_item.mail.to_discord()}",
                    )
                    if message is None:
                        logger.logger.error(
                            f"check_interia: Failed to send Interia error message for {order_item.mail}"
                        )
                        continue
                    data.interia_error_message_id_to_mail[message.id] = (
                        order_item.mail.mail
                    )
                    order_item.mail.interia_error_message_id = message.id
                    data.message_id_to_message_type[message.id] = (
                        mrowka_data.MrowkaMessageType.INTERIA_ERROR
                    )
                    await message.add_reaction(bot, "👍")
                    await mrowka_data.PisarzMrowka.write(data, safe=False)
                    continue

                if info.tracking:
                    order_item.tracking = info.tracking
                    # === AUTO UPDATE PZ: wpisz tracking do Uwag i NumerPrzesylki w Subiekcie ===
                    tracking_result = await _subiekt_post(
                        "/api/pz/update-tracking",
                        {"orderName": order_item.name, "tracking": info.tracking},
                        timeout=30,
                    )
                    if tracking_result:
                        logger.logger.info(
                            "✅ Tracking %s → PZ zaktualizowany dla %s (docId=%s)",
                            info.tracking, order_item.name, tracking_result.get("docId")
                        )
                    else:
                        logger.logger.warning(
                            "⚠️ Tracking %s znaleziony ale update PZ nie powiódł się dla %s",
                            info.tracking, order_item.name
                        )
                if info.delivery_date:
                    order_item.delivery_date = info.delivery_date
                if info.order_number:
                    order_item.order_number = info.order_number
                if info.name_surname:
                    order_item.name_surname = info.name_surname

                if (
                    order_item.tracking is not None
                    and mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_WYSLANE
                    in order_item.history.get_status().status.next_statuses()
                ):
                    await order_item.change_status(
                        bot,
                        mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_WYSLANE,
                        order_item.history.get_status().user,
                        data,
                    )
                    data.tracking_to_order_item_name[order_item.tracking] = (
                        order_item.name
                    )
                elif any(
                    [
                        info.tracking is not None,
                        info.delivery_date is not None,
                        info.order_number is not None,
                        info.name_surname is not None,
                    ]
                ):
                    await order_item.discord_update(bot, data)

                await mrowka_data.PisarzMrowka.write(data, safe=False)


# ─── RUSH MODE: przyspieszony polling po wykryciu dostawy ────────────────────

# Stan modułowy (reset przy restarcie bota — akceptowalne, maile są UNSEEN)
_rush_mode_date: Optional[str] = None        # data ostatniego rush (YYYY-MM-DD)
_rush_mode_running: bool = False             # czy coroutine juz dziala
_rush_session_oi_names: set[str] = set()    # OI które dotarły w tej sesji


def _rush_register_delivery(oi_name: str) -> None:
    """Rejestruje OI jako dostarczone w bieżącej sesji rush."""
    _rush_session_oi_names.add(oi_name)


def _maybe_start_rush_mode(bot: commands.Bot) -> None:
    """
    Triggeruje rush mode jeśli nie był jeszcze uruchomiony dzisiaj.
    Wywołać gdy wykryto pierwszy mail dostawy.
    """
    global _rush_mode_date, _rush_mode_running
    import datetime as _dt
    today = _dt.date.today().isoformat()
    if _rush_mode_running or _rush_mode_date == today:
        return
    _rush_mode_date = today
    _rush_mode_running = True
    logger.logger.info("delivery_rush_mode: startuje (3×15 min)")
    asyncio.create_task(_delivery_rush_mode(bot))


async def _delivery_rush_mode(bot: commands.Bot) -> None:
    """
    Coroutine: co 15 minut (×3) ponawia skan maili dostawczych.
    Po 3. skanowaniu generuje raport XLSX → #zalando-dostawy.
    """
    global _rush_mode_running
    try:
        for i in range(3):
            await asyncio.sleep(15 * 60)   # 15 minut
            logger.logger.info("delivery_rush_mode: poll %d/3", i + 1)
            await check_gmail_delivery(bot)
            await check_gmail_delivery_confirmed(bot)

        logger.logger.info("delivery_rush_mode: 3/3 done → generuję raport")
        await _generate_delivery_report(bot)
    except Exception as e:
        logger.logger.exception("delivery_rush_mode: %s", e)
    finally:
        _rush_mode_running = False


async def _generate_delivery_report(bot: commands.Bot) -> None:
    """
    Generuje raport XLSX + wiadomość Discord do #zalando-dostawy.

    Dla każdego ticketu gdzie OI z rush sesji:
      - PRZEPAKOWYWANIE = WSZYSTKIE non-anulowane OI są ZREALIZOWANE
      - ODŁOŻONA = przynajmniej jeden OI jeszcze nie ZREALIZOWANY
      - Lista PZ: dzisiejsze + wcześniej zrealizowane (do zebrania)
      - ZK: numer + lista PZ
    """
    try:
        import openpyxl
        import datetime as _dt
        import pathlib as _pathlib
        import io

        data = await mrowka_data.PisarzMrowka.read()
        zrealizowane_st = mrowka_data.MrowkaOrderItemStatus.ZAMOWIENIE_ZOSTALO_ZREALIZOWANE
        anulowane_st    = mrowka_data.MrowkaOrderItemStatusStatus.ANULOWANE

        # Zbierz tickety dotknięte sessją
        touched_tickets: dict[str, mrowka_data.MrowkaTicket] = {}
        for ticket in data.tickets.values():
            for oi in ticket.divided_orders.values():
                if oi.name in _rush_session_oi_names:
                    touched_tickets[ticket.name] = ticket
                    break

        if not touched_tickets:
            logger.logger.info("_generate_delivery_report: brak dotkniętych ticketów — pomijam")
            return

        # ── Buduj Excel ────────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dostawy"

        # Style
        from openpyxl.styles import Font, PatternFill, Alignment
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill_green  = PatternFill("solid", fgColor="1F6F3F")
        hdr_fill_orange = PatternFill("solid", fgColor="CC6600")
        przepak_fill = PatternFill("solid", fgColor="D6E4BC")
        odloz_fill   = PatternFill("solid", fgColor="FCE4D6")

        ws.append(["Ticket", "ZK", "Akcja", "Konto (mail)", "Tracking", "PZ sygnatura", "Status OI", "Uwagi"])
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill_green
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 28
        ws.column_dimensions["F"].width = 16
        ws.column_dimensions["G"].width = 24
        ws.column_dimensions["H"].width = 30

        # ── Discord text ───────────────────────────────────────────────────────
        discord_lines: list[str] = []
        today_str = _dt.date.today().strftime("%d.%m.%Y")
        discord_lines.append(f"📦 **Raport dostaw {today_str}** (po 45 min monitorowania)\n")

        for ticket_name, ticket in sorted(touched_tickets.items()):
            active_ois = [
                oi for oi in ticket.divided_orders.values()
                if oi.history.get_status().status.get_status() != anulowane_st
            ]
            all_done = all(
                oi.history.get_status().status == zrealizowane_st
                for oi in active_ois
            )
            akcja = "✅ PRZEPAKOWYWANIE" if all_done else "📦 ODKŁADAMY"
            row_fill = przepak_fill if all_done else odloz_fill
            zk = ticket.zk_number or "—"

            discord_lines.append(f"─────────────────────────────")
            discord_lines.append(f"🎫 **{ticket_name}**  |  ZK: `{zk}`  →  {akcja}")

            # Dzisiejsze dostawy (w sesji rush)
            today_ois = [oi for oi in active_ois if oi.name in _rush_session_oi_names]
            if today_ois:
                discord_lines.append("  📬 **Dzisiaj dotarły:**")
            for oi in today_ois:
                mail_str = oi.mail.mail if oi.mail else "—"
                pz = oi.pz_sygnatura or "—"
                track = oi.tracking or "—"
                status_name = oi.history.get_status().status.name
                discord_lines.append(f"    • `{mail_str}` | tracking: `{track}` | PZ: `{pz}`")
                ws.append([ticket_name, zk, akcja.replace("✅ ", "").replace("⏳ ", ""), mail_str, track, pz, status_name, "DZISIAJ"])
                for cell in ws[ws.max_row]:
                    cell.fill = row_fill

            # Wcześniej zrealizowane (do zebrania razem)
            earlier_ois = [
                oi for oi in active_ois
                if oi.name not in _rush_session_oi_names
                and oi.history.get_status().status == zrealizowane_st
            ]
            if earlier_ois:
                discord_lines.append("  🗃️ **Do zebrania (wcześniej zrealizowane):**")
            for oi in earlier_ois:
                mail_str = oi.mail.mail if oi.mail else "—"
                pz = oi.pz_sygnatura or "—"
                track = oi.tracking or "—"
                discord_lines.append(f"    • `{mail_str}` | PZ: `{pz}`")
                ws.append([ticket_name, zk, "", mail_str, track, pz, "ZREALIZOWANE", "WCZESNIEJ"])
                for cell in ws[ws.max_row]:
                    cell.fill = row_fill

            # Brakujące (jeszcze czekają)
            waiting_ois = [
                oi for oi in active_ois
                if oi.name not in _rush_session_oi_names
                and oi.history.get_status().status != zrealizowane_st
            ]
            if waiting_ois:
                discord_lines.append("  ❌ **Brakujące (jeszcze nie zrealizowane):**")
            for oi in waiting_ois:
                mail_str = oi.mail.mail if oi.mail else "—"
                status_name = oi.history.get_status().status.name
                discord_lines.append(f"    • `{mail_str}` | status: `{status_name}`")
                ws.append([ticket_name, zk, "", mail_str, "—", "—", status_name, "CZEKA"])

        discord_lines.append("")

        # ── Zapisz XLSX do bufora ──────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        xlsx_name = f"dostawy_{_dt.date.today().isoformat()}.xlsx"
        discord_file = discord.File(buf, filename=xlsx_name)

        # ── Wyślij na #zalando-dostawy ─────────────────────────────────────────
        channel = await dc.CHANNEL_DOSTAWY(bot)
        await channel.send(bot, content="\n".join(discord_lines), file=discord_file)

        # ── Osobna wiadomość: ZK do przepakowania + PZ do zebrania ────────────
        przepak_tickets = [
            (t_name, t) for t_name, t in sorted(touched_tickets.items())
            if all(
                oi.history.get_status().status == zrealizowane_st
                for oi in t.divided_orders.values()
                if oi.history.get_status().status.get_status() != anulowane_st
            )
        ]
        odkladamy_tickets = [
            (t_name, t) for t_name, t in sorted(touched_tickets.items())
            if (t_name, t) not in przepak_tickets
        ]

        zk_lines: list[str] = []
        if przepak_tickets:
            zk_lines.append(f"📋 **Dziś przepakowujemy {len(przepak_tickets)} ZK:**\n")
            for t_name, t in przepak_tickets:
                zk = t.zk_number or "—"
                all_pz = [
                    oi.pz_sygnatura for oi in t.divided_orders.values()
                    if oi.pz_sygnatura
                ]
                pz_str = ", ".join(f"`{p}`" for p in all_pz) if all_pz else "brak PZ"
                zk_lines.append(f"✅ **{zk}** ({t_name})")
                zk_lines.append(f"   📂 PZ do zebrania: {pz_str}")
                zk_lines.append("")

        if odkladamy_tickets:
            zk_lines.append(f"📦 **Odkładamy (niekompletne) — {len(odkladamy_tickets)} ZK:**\n")
            for t_name, t in odkladamy_tickets:
                zk = t.zk_number or "—"
                waiting = [
                    oi.mail.mail if oi.mail else oi.name
                    for oi in t.divided_orders.values()
                    if oi.history.get_status().status != zrealizowane_st
                    and oi.history.get_status().status.get_status() != anulowane_st
                ]
                zk_lines.append(f"📦 **{zk}** ({t_name})")
                zk_lines.append(f"   ⏳ Brakuje: {', '.join(waiting) or '?'}")
                zk_lines.append("")

        if zk_lines:
            await channel.send(bot, content="\n".join(zk_lines))

        logger.logger.info("_generate_delivery_report: raport wysłany do #zalando-dostawy")

    except Exception as e:
        logger.logger.exception("_generate_delivery_report: %s", e)



@logger.async_try_log()
async def update_everything(bot: commands.Bot):
    async with mrowka_data.PisarzMrowka.lock:
        data = await mrowka_data.PisarzMrowka.read(safe=False)
        ticket_to_order_items = {
            ticket.name: list(ticket.divided_orders.keys())
            for ticket in data.tickets.values()
        }

    for ticket_name in ticket_to_order_items.keys():
        for order_item_name in ticket_to_order_items[ticket_name]:
            async with mrowka_data.PisarzMrowka.lock:
                data = await mrowka_data.PisarzMrowka.read(safe=False)
                ticket = data.tickets.get(ticket_name, None)
                if ticket is None:
                    continue
                order_item = ticket.divided_orders.get(order_item_name, None)
                if order_item is None:
                    continue

                await order_item.discord_update(bot, data)
                await mrowka_data.PisarzMrowka.write(data, safe=False)


@logger.try_log(None)
def info_find_order_item_from_text(
    data: mrowka_data.MrowkaData,
    text: str,
) -> Optional[mrowka_data.MrowkaOrderItem]:
    for ticket in data.tickets.values():
        for order_item in ticket.divided_orders.values():
            if order_item.name == text:
                return order_item
            if order_item.mail:
                if order_item.mail.mail == text:
                    return order_item
                if order_item.mail.code == text:
                    return order_item
                if order_item.mail.zalando_pass == text:
                    return order_item
                if order_item.mail.interia_pass == text:
                    return order_item
            if order_item.tracking == text:
                return order_item
            if order_item.order_number == text:
                return order_item
    return None



async def create_zk_for_ticket(ticket: "mrowka_data.MrowkaTicket") -> str:
    """
    Tworzy ZK (Zamówienie od Klienta) w Subiekcie dla całego ticketu.
    Grupuje produkty po EAN (sumuje ilości).
    Klient: 'zmien_nazwe', waluta: EUR.
    Zwraca numer dokumentu ZK (sygnaturę).
    """
    # Zbierz wszystkie buty z wszystkich order_items ticketu
    all_shoes: list[mrowka_data.MrowkaShoeInfo] = []
    for oi in ticket.divided_orders.values():
        all_shoes.extend(oi.shoes)

    if not all_shoes:
        raise ValueError(f"Ticket {ticket.name} nie ma żadnych butów")

    # Pobierz EAN-y dla wszystkich linków
    all_links = list({shoe.link for shoe in all_shoes})
    link_to_entries: dict[str, list] = {}
    for link in all_links:
        link_to_entries[link] = await ean_db.get_eans_for_link(link)

    def _make_symbol(sku: str, size: str) -> str:
        """Odwzoruj sposób tworzenia Symbol jaki stosuje EnsureProductExists w C#."""
        sym = f"{sku}-{size}".replace(",", ".").replace(" ", "").replace("/", "-")
        return sym[:20]

    # Grupuj po EAN (suma ilości), zapisz też Symbol
    ean_items: dict[str, dict] = {}  # ean -> {qty, symbol}
    for shoe in all_shoes:
        entries = link_to_entries.get(shoe.link, [])
        for size, qty in shoe.size_to_quantity.items():
            if qty <= 0:
                continue
            entry = next((e for e in entries if e.size == size), None)
            if entry is None or not entry.ean:
                continue
            ean = entry.ean
            if ean not in ean_items:
                symbol = _make_symbol(entry.sku or "", entry.size or size)
                ean_items[ean] = {"qty": 0, "symbol": symbol}
            ean_items[ean]["qty"] += qty

    if not ean_items:
        raise ValueError(f"Brak EAN-ów dla ticketu {ticket.name}")

    # Zbuduj listę pozycji ZK — wysyłaj zarówno EAN jak i Symbol
    items = [
        {"EAN": ean, "Symbol": data["symbol"], "Quantity": data["qty"]}
        for ean, data in ean_items.items()
    ]

    payload = {
        "CustomerName": "zmien_nazwe",
        "Currency": "EUR",
        "TicketName": ticket.name,
        "Items": items,
    }
    result = await _sfera_cli("CreateZK", {"ZkData": payload}, timeout=120)
    if not result.get("Success"):
        raise Exception(f"CreateZK failed: {result.get('Message', '?')}")

    return result.get("DocumentNumber", "?")
